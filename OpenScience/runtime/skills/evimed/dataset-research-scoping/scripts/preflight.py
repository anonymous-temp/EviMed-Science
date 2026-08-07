#!/usr/bin/env python3
"""Deterministic preflight for dataset research scoping deliverables.

Blocks on what a reader cannot check for themselves:

1. an identifier copied out of the source data,
2. profile numbers that do not match the script that claims to produce them,
3. an infeasible verdict that does not name the field it is missing, does not
   show that the gap actually binds, or does not say what the data can still
   answer,
4. a post-hoc power calculation,
5. an evidence base below the Phase 3 floors — too few works, too few channels,
   too few full texts, citations a reader cannot open, or a work cited in the
   report that is absent from the map,
6. a surviving question with no novelty statement,
7. an analysis that never left the ground — fewer than six of the eight analysis
   families considered anywhere, a surviving question with no named estimator, or
   not one internal-consistency identity run against the schema.

Warns on four a reader can see: an unqualified fill rate, an unexplained blank
in the seven-element matrix, an external resource with no join key, and a
preprocessing item with no stated consequence.

It does not check how many candidate questions were produced. The right number
depends on the data, and a quota would only teach runs to pad. It does check how
much of the field was read before deciding, because that number does not depend
on the data at all: a run searched one index, cited twelve works, and returned a
portfolio whose directions nobody could tell were new.

Evidence breadth is only half of it. A later run cleared every breadth floor —
49 works, nine channels, six full texts — and still handed back two descriptive
audits, because it never asked whether prediction, class-level comparison,
association mining, or multi-library synthesis were on the table, and deleted the
metabolic-effect question for want of laboratory parameters while 22 longitudinal
weight records sat in the vitals table. Depth is checked separately for that
reason.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from evidence_floor import (  # noqa: E402  - resolved from this script's own directory
    EVIDENCE_MAP,
    check_evidence_breadth,
    check_novelty_statements,
    read_text,
)

REQUIRED_OUTPUTS = (
    "data-profile.md",
    "data-profile.py",
    "data-quality.md",
    EVIDENCE_MAP,
    "feasibility-matrix.md",
    "external-linkage.md",
    "research-portfolio.md",
    "study-protocol.md",
    "scoping-run.json",
)
PROSE_OUTPUTS = (
    "data-profile.md",
    "data-quality.md",
    EVIDENCE_MAP,
    "feasibility-matrix.md",
    "external-linkage.md",
    "research-portfolio.md",
    "study-protocol.md",
)
# Columns whose values identify a person or an episode of care. Their values may
# never leave the data; a run refers to subjects by pseudonyms it assigns.
# Kept identical to profile_dataset.py's rule. Matching the subject word alone
# swept in RECORD_DATE and RECORD_CONTENT, whose values are timestamps and
# clinical text: every date in the report would then have read as a leaked
# identifier and blocked a sound package.
# Matching the subject word as a substring was too broad (RECORD_DATE), and
# matching it as a whole word was too narrow: a real front-page column is
# MED_REC_NO — the medical record number, abbreviated — which neither rule
# caught, so its full vocabulary of hospital numbers was printed into the
# profile and handed over as a clean deliverable. Names are split into tokens
# and a column identifies when a subject token meets an id-shaped last token.
IDENTIFIER_EXPLICIT = re.compile(r"^(?:id|mrn|姓名|患者姓名|身份证号?|病案号|住院号|门诊号|就诊号)$", re.I)
IDENTIFIER_SUBJECT_TOKENS = frozenset({
    "patient", "subject", "person", "case", "record", "rec", "admission",
    "visit", "encounter", "inpatient", "outpatient", "mrn",
})
# "adm" and "reg" were in this set and masked ADM_DEPT_CODE, the admitting
# department — a covariate, not a person. A subject token must be a word that
# names the subject, not any abbreviation that begins one.
IDENTIFIER_SUFFIX_TOKENS = frozenset({"id", "no", "num", "number", "code", "sn"})
IDENTIFIER_SUBJECT_CJK = re.compile(r"(?:病案|住院|门诊|患者|病人|就诊|身份证|姓名)")


def is_identifying(name: str) -> bool:
    """True when a column's values identify a person or an episode of care."""
    cleaned = name.strip()
    if IDENTIFIER_EXPLICIT.match(cleaned):
        return True
    tokens = [t for t in re.split(r"[^0-9A-Za-z]+", cleaned) if t]
    if tokens and tokens[-1].lower() in IDENTIFIER_SUFFIX_TOKENS:
        if any(t.lower() in IDENTIFIER_SUBJECT_TOKENS for t in tokens):
            return True
        if IDENTIFIER_SUBJECT_CJK.search(cleaned):
            return True
    return False


# An identifier short enough to collide with an ordinary number in prose is not
# evidence of leakage; below this length a match is not reported.
IDENTIFIER_MIN_LENGTH = 5
POST_HOC_POWER = re.compile(
    r"(?:事后功效|事后检验效能|观测功效|观察到的功效|post[\s-]?hoc power|observed power|retrospective power)",
    re.I,
)
INFEASIBLE_LINE = re.compile(r"(?:不可行|infeasible|not feasible)", re.I)
MISSING_FIELD_LINE = re.compile(r"(?:缺失字段|缺少字段|缺字段|missing field|missing column)", re.I)
# Naming the missing field was the whole bar, and it turned out to reward
# refusal: a run could discharge its duty by pointing at an absent column and
# never ask what the data could still answer. An infeasible verdict now also
# owes the strongest remaining question, or an explicit statement that none
# exists — stated somewhere in the matrix, not necessarily on the verdict line.
FALLBACK_MENTION = re.compile(
    r"(?:退而求其次|降级|替代设计|仍可|仍然可以|还能|可改为|可降级为|弱化为|"
    r"fallback|degraded design|can still|weaker question|instead answer)",
    re.I,
)
# A binding test: the gap was compared against something in the same units
# rather than asserted from a general rule about what a method requires.
BINDING_TEST = re.compile(
    r"(?:半衰期|给药间隔|变异(?:度|系数)|CV|倍|数量级|相对极差|敏感性分析|"
    r"half-life|dosing interval|fold|order of magnitude|sensitivity analys)",
    re.I,
)
VERDICT_MENTION = re.compile(r"(?:判定|裁定|verdict)", re.I)
FEASIBLE_MENTION = re.compile(r"(?:可行|feasible)", re.I)
# Phase 4b. A run that never asks whether prediction is on the table does not
# produce a prediction question, and the report then reads as though the data
# could not carry one. The run this exists for deleted the metabolic-effect
# question for want of lab parameters while 22 weight records sat in the vitals
# table, and never considered prediction, association mining, or causal
# inference at all — its two survivors were both descriptive audits.
ANALYSIS_FAMILIES = (
    ("prediction", re.compile(r"预测|prediction|predictive model|建模", re.I)),
    ("class-level", re.compile(r"类药|药物类|class[- ]level|同类药|drug class", re.I)),
    ("association", re.compile(r"关联挖掘|关联规则|association (?:rule|mining)|共现|market[- ]basket", re.I)),
    ("causal", re.compile(r"因果|causal|target trial|目标试验|IPTW|g-formula|反事实", re.I)),
    ("pharmacovigilance", re.compile(r"不良反应|药物警戒|pharmacovigilance|ADR|信号检测|ROR|FAERS", re.I)),
    ("external-linkage", re.compile(r"外部(?:资源|链接|接驳)|external linkage|公共数据库|RxNorm|LOINC|ATC", re.I)),
    ("multi-library", re.compile(r"多库|multi[- ]librar|跨库|证据合成|pooled|meta[- ]analy", re.I)),
    ("descriptive", re.compile(r"描述性|审计|descriptive|audit", re.I)),
)
MIN_FAMILIES_CONSIDERED = 6
# A surviving question without an estimator is a topic, not a design.
ESTIMATOR_MENTION = re.compile(
    r"估计量|估计器|estimator|回归|regression|logistic|cox|混合效应|mixed[- ]effect|"
    r"随机森林|random forest|lasso|贝叶斯|bayes|bootstrap|交叉验证|cross[- ]valid|"
    r"IPTW|g-formula|倾向|propensity|ROR|PRR|apriori|FP[- ]growth|置信区间|confidence interval",
    re.I,
)
# An identity the schema implies, actually run, with a count. The TDM extract's
# total = parent + metabolite held in 6 of 6 sets, which is what licensed using
# any of the three numbers; no run had ever checked it.
# A topic named after the activity performed rather than the question answered.
# "TDM 采样实践审计" was rejected outright by the researcher — "这是学术语言吗" —
# and the skill had put the word there itself. The register is part of the design:
# a compliance word in a title tells a reviewer the work is administrative before
# they read a line. Checked on headings only, so the same word may still be
# discussed in the prose.
ACTIVITY_TITLE = re.compile(
    r"^#{2,4}[^\n]*(?:审计|梳理|摸底|盘点|情况分析|现状调查|\baudit\b|\bstocktak)",
    re.I | re.M,
)
IDENTITY_CHECK = re.compile(
    r"(?:一致性|恒等|校验|identity|consistency check|internal check|对账)[^\n]{0,80}?\d",
    re.I,
)
FILL_RATE_MENTION = re.compile(r"(?:填充率|fill rate|completeness)", re.I)
COMPLETENESS_QUALIFIER = re.compile(r"(?:density|documentation|breadth|predictive|Weiskopf)", re.I)
SEVEN_ELEMENTS = (
    ("eligibility", re.compile(r"(?:入组|合格|纳入)标准|eligibility", re.I)),
    ("treatment", re.compile(r"治疗策略|干预策略|treatment strateg", re.I)),
    ("assignment", re.compile(r"分配机制|assignment procedure", re.I)),
    ("timeZero", re.compile(r"时间零点|随访起[点始]|time zero|follow[- ]up start", re.I)),
    ("outcome", re.compile(r"结局|outcome", re.I)),
    ("contrast", re.compile(r"因果对比|causal contrast", re.I)),
    ("analysisPlan", re.compile(r"分析计划|analysis plan", re.I)),
)
BLANK_CELL = re.compile(r"(?:❌|⚠️|待定|TBD|N/?A|——|—)\s*$")
JOIN_KEY_MENTION = re.compile(r"(?:连接键|关联键|连接字段|join key|linkage key|join on|通过\s*`?\w+`?\s*连接)", re.I)
CONSEQUENCE_MENTION = re.compile(r"(?:否则|不做|将导致|会导致|后果|otherwise|leads to|breaks|invalidat)", re.I)


def identifier_values(dataset_paths: list[Path]) -> set[str]:
    """Every value held by an identifier-shaped column of the source data."""
    values: set[str] = set()
    for path in dataset_paths:
        suffix = path.suffix.lower()
        tables: list[tuple[list[str], list[list[str]]]] = []
        if suffix in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError:  # pragma: no cover - environment dependent
                continue
            workbook = load_workbook(path, read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                rows = [["" if c is None else str(c) for c in row] for row in sheet.iter_rows(values_only=True)]
                if rows:
                    tables.append((rows[0], rows[1:]))
            workbook.close()
        else:
            delimiter = "\t" if suffix in {".tsv", ".tab"} else ","
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle, delimiter=delimiter))
            if rows:
                tables.append(([str(h) for h in rows[0]], [[str(c) for c in r] for r in rows[1:]]))

        for header, rows in tables:
            for index, column_name in enumerate(header):
                if not is_identifying(str(column_name)):
                    continue
                for row in rows:
                    if index >= len(row):
                        continue
                    value = str(row[index]).strip()
                    if len(value) >= IDENTIFIER_MIN_LENGTH:
                        values.add(value)
    return values


def scannable_files(root: Path, dataset_paths: list[Path]) -> list[Path]:
    """Every text file the run leaves behind, except the source data itself.

    Scanning only the declared deliverables was not enough. A run wrote its own
    working file holding {"pseudonyms": {"900004": "P1", ...}} — the mapping back
    to real hospital numbers, which defeats the pseudonyms entirely — and because
    that file was not on the declared list it was never looked at. Anything left
    in the workspace can be read by whoever receives it.
    """
    sources = {path.resolve() for path in dataset_paths}
    files = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.is_symlink():
            continue
        if path.resolve() in sources:
            continue
        # Retrieved source documents are quoted evidence, not run output.
        if ".evimed-sources" in path.parts:
            continue
        if path.suffix.lower() in {".xlsx", ".xlsm", ".xls", ".png", ".jpg", ".pdf", ".zip", ".gz"}:
            continue
        files.append(path)
    return files


def check_identifier_leakage(root: Path, dataset_paths: list[Path], issues: list[str]) -> int:
    values = identifier_values(dataset_paths)
    if not values:
        return 0
    leaked = 0
    for path in scannable_files(root, dataset_paths):
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        for value in sorted(values):
            if value in text:
                leaked += 1
                name = path.relative_to(root).as_posix()
                issues.append(
                    f"{name}: carries the source identifier {value!r}. Refer to subjects by a pseudonym you "
                    "assign, and never write the mapping back to the source value into the workspace."
                )
    return leaked


def check_profile_recomputable(root: Path, dataset_paths: list[Path], issues: list[str]) -> bool:
    script = root / "data-profile.py"
    recorded = root / "data-profile.json"
    if not script.is_file():
        issues.append("data-profile.py is missing: the profile numbers must come from a script that is kept.")
        return False
    if not recorded.is_file():
        issues.append("data-profile.json is missing: data-profile.py must write the machine-readable profile it renders.")
        return False
    if not dataset_paths:
        issues.append(
            "scoping-run.json: priorDataContact.filesReceived names no readable dataset file, "
            "so the profile cannot be recomputed."
        )
        return False
    with tempfile.TemporaryDirectory() as work:
        target = Path(work) / "recomputed.json"
        try:
            completed = subprocess.run(
                [sys.executable, str(script), *[str(p) for p in dataset_paths],
                 "--json", str(target), "--markdown", str(Path(work) / "recomputed.md")],
                cwd=root, capture_output=True, text=True, timeout=600,
            )
        except (OSError, subprocess.SubprocessError) as error:
            issues.append(f"data-profile.py could not be re-run: {error}")
            return False
        if completed.returncode != 0:
            issues.append(f"data-profile.py exited {completed.returncode} when re-run: {completed.stderr.strip()[:400]}")
            return False
        try:
            fresh = json.loads(target.read_text(encoding="utf-8"))
            stored = json.loads(recorded.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            issues.append(f"the recomputed profile could not be compared: {error}")
            return False
    if fresh != stored:
        issues.append(
            "data-profile.json does not match what data-profile.py produces from the same inputs; "
            "the reported numbers are not the script's."
        )
        return False
    return True


def check_infeasible_verdicts(root: Path, issues: list[str]) -> int:
    text = read_text(root, "feasibility-matrix.md")
    if not text.strip():
        issues.append("feasibility-matrix.md is empty: every candidate question needs a verdict.")
        return 0
    verdicts = 0
    for line_number, line in enumerate(text.splitlines(), start=1):
        if not INFEASIBLE_LINE.search(line):
            continue
        verdicts += 1
        if not MISSING_FIELD_LINE.search(line):
            issues.append(
                f"feasibility-matrix.md line {line_number}: an infeasible verdict must name the missing field."
            )
    if verdicts:
        # Whole-document checks: a refusal is only a verdict once it has been
        # shown to bind and the remaining question has been stated. Requiring
        # these per line would force boilerplate onto every row, so they are
        # required of the matrix as a whole.
        if not BINDING_TEST.search(text):
            issues.append(
                "feasibility-matrix.md declares questions infeasible without showing anywhere that the gap "
                "actually binds — compare it against the effect being studied in the same units "
                "(half-life against dosing interval, assay CV against between-subject spread), "
                "rather than citing a general requirement."
            )
        if not FALLBACK_MENTION.search(text):
            issues.append(
                "feasibility-matrix.md declares questions infeasible without naming the strongest question "
                "the data can still answer. State the degraded design and its assumption, or say explicitly "
                "that none exists."
            )
    return verdicts


def check_breadth_and_novelty(root: Path, issues: list[str]) -> dict:
    """The Phase 3 floors, plus one novelty statement per surviving question."""
    metrics = check_evidence_breadth(root, PROSE_OUTPUTS, issues)
    matrix = read_text(root, "feasibility-matrix.md")
    feasible = 0
    for line in matrix.splitlines():
        if INFEASIBLE_LINE.search(line):
            continue
        if VERDICT_MENTION.search(line) and FEASIBLE_MENTION.search(line):
            feasible += 1
    metrics["feasibleQuestions"] = feasible
    metrics.update(check_novelty_statements(root, "research-portfolio.md", feasible, issues))
    metrics.update(check_analytical_depth(root, feasible, issues))
    return metrics


def check_analytical_depth(root: Path, feasible: int, issues: list[str]) -> dict:
    """Whether designs were produced, or only verdicts about designs.

    Evidence breadth is already checked; this is the other half. A package can
    cite fifty works, draw on nine channels, and still hand back two descriptive
    audits because it never asked whether prediction or causal inference were on
    the table, never ran the identities that license its own numbers, and never
    named an estimator for anything that survived.
    """
    matrix = read_text(root, "feasibility-matrix.md")
    portfolio = read_text(root, "research-portfolio.md")
    protocol = read_text(root, "study-protocol.md")
    quality = read_text(root, "data-quality.md")
    profile = read_text(root, "data-profile.md")
    considered = sorted(
        name for name, pattern in ANALYSIS_FAMILIES if pattern.search(matrix) or pattern.search(portfolio)
    )
    if len(considered) < MIN_FAMILIES_CONSIDERED:
        missing = [name for name, _ in ANALYSIS_FAMILIES if name not in considered]
        issues.append(
            f"only {len(considered)} of {len(ANALYSIS_FAMILIES)} analysis families appear anywhere in the "
            f"matrix or the portfolio ({', '.join(considered) or 'none'}); the floor is "
            f"{MIN_FAMILIES_CONSIDERED}. Never considered: {', '.join(missing)}. Walk the Phase 4b list "
            "and record a line for each — which fields it would consume, what it would answer, and only "
            "if it truly is, why it is off the table."
        )
    estimators = len(ESTIMATOR_MENTION.findall(portfolio)) + len(ESTIMATOR_MENTION.findall(protocol))
    if feasible > 0 and estimators < feasible:
        issues.append(
            f"{estimators} estimator mentions across research-portfolio.md and study-protocol.md for "
            f"{feasible} surviving questions. A question without a named estimator is a topic, not a "
            "design — say what is being estimated and with what."
        )
    activity_titles = ACTIVITY_TITLE.findall(portfolio) + ACTIVITY_TITLE.findall(protocol)
    if activity_titles:
        issues.append(
            f"{len(activity_titles)} topic heading(s) are named after the activity performed rather "
            "than the question answered — 审计 / 梳理 / 摸底 / 盘点 / audit and the like. Name a topic "
            "after what the reader learns or what is estimated: not 「采样实践审计」 but 「常规 TDM "
            "记录中稳态达标率的估计及其与群体分布的偏离」."
        )
    identities = len(IDENTITY_CHECK.findall(quality)) + len(IDENTITY_CHECK.findall(profile))
    if identities == 0:
        issues.append(
            "no internal-consistency identity is reported with a count in data-quality.md or "
            "data-profile.md. Run every identity the schema implies — a total against the sum of its "
            "parts, a length of stay against its dates, a value against the reference range in its own "
            "row — and report how many rows passed. These are what license using the numbers at all."
        )
    return {
        "analysisFamiliesConsidered": considered,
        "estimatorMentions": estimators,
        "identityChecksReported": identities,
        "activityStyleTitles": len(activity_titles),
    }


def check_post_hoc_power(root: Path, issues: list[str]) -> None:
    for name in PROSE_OUTPUTS:
        text = read_text(root, name)
        for line_number, line in enumerate(text.splitlines(), start=1):
            if POST_HOC_POWER.search(line):
                issues.append(
                    f"{name} line {line_number}: post-hoc power is uninformative for a fixed dataset. "
                    "Report the minimum detectable effect and the expected interval width instead."
                )


def check_prior_data_contact(root: Path, issues: list[str]) -> list[Path]:
    raw = read_text(root, "scoping-run.json")
    if not raw.strip():
        issues.append("scoping-run.json is missing.")
        return []
    try:
        receipt = json.loads(raw)
    except json.JSONDecodeError as error:
        issues.append(f"scoping-run.json is not valid JSON: {error}")
        return []
    contact = receipt.get("priorDataContact")
    if not isinstance(contact, dict):
        issues.append(
            "scoping-run.json: priorDataContact is required. Record it before profiling — "
            "written afterwards it records what you already knew."
        )
        return []
    files = contact.get("filesReceived")
    if not isinstance(files, list) or not files:
        issues.append("scoping-run.json: priorDataContact.filesReceived must list the files received.")
        files = []
    for field in ("partsInspected", "outcomeDistributionSeen"):
        if field not in contact:
            issues.append(f"scoping-run.json: priorDataContact.{field} is required.")

    paths = []
    for entry in files:
        if not isinstance(entry, str) or ".." in entry or entry.startswith("/"):
            issues.append(f"scoping-run.json: priorDataContact.filesReceived entry {entry!r} must be a workspace path.")
            continue
        candidate = root / entry
        if candidate.is_file():
            paths.append(candidate)
        else:
            issues.append(f"scoping-run.json: priorDataContact.filesReceived entry {entry!r} does not exist.")
    return sorted(paths, key=lambda p: p.name)


def collect_warnings(root: Path) -> list[str]:
    warnings: list[str] = []

    for name in ("data-profile.md", "data-quality.md"):
        text = read_text(root, name)
        if FILL_RATE_MENTION.search(text) and not COMPLETENESS_QUALIFIER.search(text):
            warnings.append(
                f"{name}: reports a fill rate without saying which of Weiskopf's four completeness "
                "definitions it measures."
            )

    matrix = read_text(root, "feasibility-matrix.md")
    if matrix:
        for label, pattern in SEVEN_ELEMENTS:
            for line in matrix.splitlines():
                if pattern.search(line) and BLANK_CELL.search(line.rstrip()):
                    warnings.append(
                        f"feasibility-matrix.md: the {label} element is left blank without an explanation."
                    )
                    break

    linkage = read_text(root, "external-linkage.md")
    if linkage.strip() and not JOIN_KEY_MENTION.search(linkage):
        warnings.append("external-linkage.md: names resources without naming the field that joins them.")

    quality = read_text(root, "data-quality.md")
    for line in quality.splitlines():
        stripped = line.strip()
        if not stripped.startswith(("-", "*", "|")) or len(stripped) < 8:
            continue
        if re.search(r"(?:预处理|preprocess|清洗|clean)", stripped, re.I) and not CONSEQUENCE_MENTION.search(stripped):
            warnings.append(
                "data-quality.md: a preprocessing item does not say which analysis breaks if it is skipped."
            )
            break

    return warnings


def main() -> int:
    parser = argparse.ArgumentParser(description="Preflight dataset research scoping deliverables.")
    parser.add_argument("--workspace", default=".")
    args = parser.parse_args()
    root = Path(args.workspace).resolve()

    issues: list[str] = []
    for name in REQUIRED_OUTPUTS:
        if not (root / name).is_file():
            issues.append(f"{name} is missing.")

    dataset_paths = check_prior_data_contact(root, issues)
    leaked = check_identifier_leakage(root, dataset_paths, issues)
    recomputable = check_profile_recomputable(root, dataset_paths, issues)
    infeasible = check_infeasible_verdicts(root, issues)
    check_post_hoc_power(root, issues)
    evidence = check_breadth_and_novelty(root, issues)
    warnings = collect_warnings(root)

    payload = {
        "ok": not issues,
        "workspace": str(root),
        "metrics": {
            "datasetFiles": len(dataset_paths),
            "identifierLeaks": leaked,
            "profileRecomputable": recomputable,
            "infeasibleVerdicts": infeasible,
            **evidence,
        },
        "issues": issues,
        "warnings": warnings,
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if not issues else 1


if __name__ == "__main__":
    sys.exit(main())

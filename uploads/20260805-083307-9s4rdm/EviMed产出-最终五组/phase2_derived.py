#!/usr/bin/env python3
"""Phase 2 — domain-derived quantities for the TDM schema sample.

Computes the quantities that carry information in the TDM domain and the
identity checks that license using the numbers at all. Subjects are referred
to by pseudonyms P1..P5 assigned in memory in stable order (sorted by
PATIENT_ID); the mapping is never written to disk.

Every number here is an existence proof for the pipeline on the 5-row schema
sample, NOT an estimate of the full cohort.
"""
import json
import re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

SRC = Path("/workspace/20260803TDM.xlsx")


def parse_dt(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("0/0/0 00:00:00", "0/0/0 0:00:00"):
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def rows_of(sheet):
    it = sheet.iter_rows(values_only=True)
    header = list(next(it))
    out = []
    for r in it:
        out.append(dict(zip(header, r)))
    return out


wb = load_workbook(SRC, read_only=True, data_only=True)
front = rows_of(wb["病案首页"])
orders = rows_of(wb["医嘱记录"])
labs = rows_of(wb["检验"])
diag = rows_of(wb["诊断记录"])
vitals = rows_of(wb["体征"])

# ---- pseudonyms in memory only ----
def pid_norm(v):
    try:
        return int(str(v).strip().lstrip("0") or 0)
    except Exception:
        return int(re.sub(r"\D", "", str(v)) or 0)
pids = sorted({pid_norm(r["PATIENT_ID"]) for r in front})
pseudo = {pid: f"P{i+1}" for i, pid in enumerate(pids)}

rep = {}

# ---- Identity checks ----
# 1. total = parent + metabolite for aripiprazole triples
triples = defaultdict(list)
for r in labs:
    if r["PROJECT_NAME"] in ("阿立哌唑", "脱氢阿立哌唑", "总阿立哌唑"):
        triples[(r["CASE_NO"], r["TEST_DATE"], r["SAMPLENO"])].append(r)
id_total = {"sets": 0, "passed": 0}
for key, rows in triples.items():
    by_name = {r["PROJECT_NAME"]: r["TEST_RESULT"] for r in rows}
    if set(by_name) == {"阿立哌唑", "脱氢阿立哌唑", "总阿立哌唑"}:
        id_total["sets"] += 1
        if abs(by_name["总阿立哌唑"] - (by_name["阿立哌唑"] + by_name["脱氢阿立哌唑"])) < 1e-6:
            id_total["passed"] += 1
rep["identity_total_parent_metabolite"] = id_total

# 2. DAY_TOTAL vs DIS_DATE - IN_DATE
id_los = {"rows": 0, "passed": 0}
for r in front:
    tin, tout = parse_dt(r["IN_DATE"]), parse_dt(r["DIS_DATE"])
    if tin and tout:
        id_los["rows"] += 1
        if (tout - tin).days + 1 == int(r["DAY_TOTAL"]):
            id_los["passed"] += 1
rep["identity_length_of_stay"] = id_los

# 3. TEST_RESULT vs REFFR_SCOPE (within-range check, where a range exists)
id_ref = {"rows_with_range": 0, "within": 0}
for r in labs:
    scope = str(r["REFFR_SCOPE"] or "").strip()
    m = re.match(r"^\s*([\d.]+)\s*-\s*([\d.]+)\s*$", scope)
    if m:
        lo, hi = float(m.group(1)), float(m.group(2))
        id_ref["rows_with_range"] += 1
        if lo <= float(r["TEST_RESULT"]) <= hi:
            id_ref["within"] += 1
rep["identity_reference_range"] = id_ref

# 4. vitals BP composite split
bp_composite = sum(1 for r in vitals if r["SIGN_TYPE"] == "血压" and "/" in str(r["RECORD_CONTENT"]))
rep["identity_bp_composite"] = {"bp_rows": sum(1 for r in vitals if r["SIGN_TYPE"] == "血压"), "splittable": bp_composite}

# ---- drug orders: keep only drug rows ----
drug_orders = [r for r in orders if r.get("DRUG_FLAG") == 1]
herb_rows = [r for r in drug_orders if str(r.get("MEDICATION_WAY") or "") in ("煎服", "先煎", "副药") or "J" in str(r.get("DRUG_NAME") or "")]
western = [r for r in drug_orders if r not in herb_rows]
rep["orders"] = {
    "drug_rows": len(drug_orders),
    "herb_rows": len(herb_rows),
    "western_rows": len(western),
    "discharge_takehome": sum(1 for r in drug_orders if str(r.get("MEDICATION_WAY")) == "出院带药"),
    "long_term": sum(1 for r in orders if r.get("LONG_D_NO") == 1),
    "frequency_vocab": sorted({str(r.get("FREQUENCY")) for r in drug_orders}),
}

# aripiprazole orders: identify by name containing 阿立哌唑
ari_orders = [r for r in western if "阿立哌唑" in str(r.get("DRUG_NAME") or "") + str(r.get("ORDER_CONTENT") or "")]

# daily dose decoder for aripiprazole orders (FREQUENCY coding)
def freq_per_day(f):
    f = str(f or "").strip().upper()
    if f in ("QD", "QN", "ONCE"):
        return 1
    if f == "BID":
        return 2
    if f == "TID":
        return 3
    m = re.match(r"^BID(\d+)$", f)
    if m:
        return int(m.group(1)) * 2
    m = re.match(r"^(\d+)D(\d+)$", f)  # e.g. W4D8 style
    if m:
        return 1.0 * int(m.group(2)) / int(m.group(1))
    if f == "ALWAYS":
        return 1
    if f == "QOD":
        return 0.5
    return None


# ---- exposure reconstruction: dose in force on sampling date ----
labs_by_pseudo = defaultdict(list)
for r in labs:
    labs_by_pseudo[pseudo[pid_norm(r["PATIENT_ID"])]].append(r)

cd = []
for pid in pids:
    p = pseudo[pid]
    for lab in labs_by_pseudo[p]:
        if lab["PROJECT_NAME"] != "阿立哌唑":
            continue
        sampleno = str(lab["SAMPLENO"])
        m = re.search(r"(\d{4})(\d{2})(\d{2})", sampleno)
        if not m:
            continue
        sample_date = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        dose = None
        for o in ari_orders:
            if pid_norm(o["PATIENT_ID"]) != pid:
                continue
            start, end = parse_dt(o["START_DATETIME"]), parse_dt(o["END_DATETIME"])
            if start and start <= sample_date and (end is None or end >= sample_date):
                per = freq_per_day(o["FREQUENCY"])
                if per:
                    dose = float(o["DOSAGE"]) * per
                    break
        if dose:
            cd.append({
                "subject": p,
                "sampleno_date": sample_date.strftime("%Y-%m-%d"),
                "dose_mg_day": dose,
                "conc_ng_ml": float(lab["TEST_RESULT"]),
                "cd_ng_ml_per_mg": float(lab["TEST_RESULT"]) / dose,
            })

rep["cd_ratio"] = {
    "n_observations": len(cd),
    "min": min(x["cd_ng_ml_per_mg"] for x in cd),
    "max": max(x["cd_ng_ml_per_mg"] for x in cd),
    "subjects": len({x["subject"] for x in cd}),
    "rows": cd,
}

# ---- dehydro-aripiprazole / aripiprazole ratio (CYP2D6 phenotype proxy) ----
ratio_rows = []
for key, rows in triples.items():
    by_name = {r["PROJECT_NAME"]: r["TEST_RESULT"] for r in rows}
    if "阿立哌唑" in by_name and "脱氢阿立哌唑" in by_name and by_name["阿立哌唑"] > 0:
        ratio_rows.append({"ratio": by_name["脱氢阿立哌唑"] / by_name["阿立哌唑"], "sample": str(key[2])})
rep["dha_ari_ratio"] = {
    "n": len(ratio_rows),
    "min": min(r["ratio"] for r in ratio_rows),
    "max": max(r["ratio"] for r in ratio_rows),
    "spread_fold": max(r["ratio"] for r in ratio_rows) / min(r["ratio"] for r in ratio_rows),
}

# ---- within-subject shift (subject sampled twice at same dose) ----
shift = []
by_subj_dose = defaultdict(list)
for x in cd:
    by_subj_dose[(x["subject"], x["dose_mg_day"])].append(x)
for (s, d), xs in by_subj_dose.items():
    if len(xs) >= 2:
        cs = sorted(x["conc_ng_ml"] for x in xs)
        shift.append({"subject": s, "dose": d, "conc_range": cs})
rep["within_subject_shift"] = shift

# ---- weight records ----
wt = [(pseudo[pid_norm(r["PATIENT_ID"])], r["RECORD_DATE"], r["RECORD_CONTENT"]) for r in vitals if r["SIGN_TYPE"] == "体重"]
wt_by_subj = defaultdict(list)
for s, d, v in wt:
    wt_by_subj[s].append({"date": str(d)[:10], "value": v})
rep["weight"] = {
    "total_records": len(wt),
    "subjects": len(wt_by_subj),
    "min_records_per_subject": min(len(v) for v in wt_by_subj.values()),
    "max_records_per_subject": max(len(v) for v in wt_by_subj.values()),
    "per_subject": {k: v for k, v in sorted(wt_by_subj.items())},
}

# ---- diagnoses relevant to ADR proxies ----
diag_names = "|".join(str(r.get("DIAGNOSTIC_NAME") or "") for r in diag)
other_names = "|".join(str(r.get("OTHER_DIAGNOSIS_NAME") or "") for r in front)
rep["adr_proxy_diagnoses"] = {
    "高泌乳素血症": other_names.count("高泌乳素血症"),
    "肌张力障碍": (diag_names + other_names).count("肌张力障碍"),
    "迟发性运动障碍": (diag_names + other_names).count("迟发性运动障碍"),
    "肝功能不全": other_names.count("肝功能不全"),
    "TCM_syndrome_rows": sum(1 for r in diag if "证" in str(r.get("DIAGNOSTIC_NAME") or "")),
}

# ---- dose change after TDM result (G3 feasibility probe) ----
ari_orders_dated = []
for o in ari_orders:
    s, e = parse_dt(o["START_DATETIME"]), parse_dt(o["END_DATETIME"])
    if s:
        ari_orders_dated.append({"start": s, "end": e, "dose": float(o["DOSAGE"]) * (freq_per_day(o["FREQUENCY"]) or 0)})
dose_changes_after_test = 0
for lab in labs:
    if lab["PROJECT_NAME"] != "阿立哌唑":
        continue
    t = parse_dt(lab["TEST_DATE"])
    if not t:
        continue
    for o in ari_orders_dated:
        if o["start"] > t + timedelta(days=7):
            dose_changes_after_test += 1
            break
rep["dose_change_after_test_within_7d"] = dose_changes_after_test

print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))

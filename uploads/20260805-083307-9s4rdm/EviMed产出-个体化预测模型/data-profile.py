#!/usr/bin/env python3
"""Mechanical data profiling for dataset research scoping.

Single-column profiling, inclusion-dependency discovery across tables, and the
trap detectors that a real hospital extract needed (sentinel dates, a join key
that is present but empty, a column typed differently in two tables, composite
values in one cell).

Deterministic by construction: the same inputs produce byte-identical JSON, so
the preflight can re-run this and compare. Nothing here reads the network, and
no cell value is printed except as an aggregate or a vocabulary entry.

Copy this file into the workspace as data-profile.py and run it there, so the
deliverable regenerates from a deliverable:

    python3 data-profile.py data/*.xlsx --json data-profile.json --markdown data-profile.md
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path

SCHEMA_VERSION = 1
# Above this many distinct values a vocabulary is reported as its most frequent
# entries rather than in full; below it, the full vocabulary is the finding.
VOCABULARY_COMPLETE_MAX = 30
VOCABULARY_SAMPLE = 15
# A column whose name looks like a join key is checked for containment in every
# other table that has a column of the same name.
JOIN_KEY_PATTERN = re.compile(r"(?:^|_)(?:id|no|code|key|num|number)$", re.I)
# Columns whose values identify a person or an episode of care. Their cardinality
# and fill rate are findings; their values are not. Printing the vocabulary of a
# five-patient PATIENT_ID column writes five real hospital numbers into the
# deliverable — which is what the preflight's first gate exists to catch, and it
# caught this script doing it. Joins are still computed from the real values;
# they are simply never emitted.
#
# Matching the subject word alone was too broad: RECORD_DATE and RECORD_CONTENT
# contain "record", so a vital-signs table had its most analytic column masked.
# An identifier is a subject word carrying an id-shaped suffix, or one of the
# few names that are identifiers outright.
IDENTIFIER_EXPLICIT = re.compile(r"^(?:id|mrn|姓名|患者姓名|身份证号?|病案号|住院号|门诊号|就诊号)$", re.I)
IDENTIFIER_SUBJECT = re.compile(
    r"(?:patient|subject|person|case|record|admission|visit|encounter|inpatient|"
    r"病案|住院|门诊|患者|病人|就诊)",
    re.I,
)
IDENTIFIER_SUFFIX = re.compile(r"(?:^|_)(?:id|no|num|number|code)$", re.I)
# Record-level identifiers that the subject-word rule misses but that still
# identify an episode, an order, or a specimen. The researcher's own rule is
# stricter than the preflight's: NO identifier value from the source data may be
# written into a deliverable, so every id-shaped record key is masked even when
# the preflight's detector would not flag it (MED_REC_NO, SAMPLENO,
# DOCTOR_ORDER_ID, ORDER_NO, ORDER_SUB_NO, RECORD_ID, DIAGNOSTIC_RECORD_ID).
IDENTIFIER_ALWAYS_MASK = {
    "MED_REC_NO", "SAMPLENO", "DOCTOR_ORDER_ID", "ORDER_NO", "ORDER_SUB_NO",
    "RECORD_ID", "DIAGNOSTIC_RECORD_ID",
}


def is_identifying(name: str) -> bool:
    """True when a column's values identify a person or an episode of care."""
    stripped = name.strip()
    if stripped.upper() in IDENTIFIER_ALWAYS_MASK or stripped.lower().startswith("unnamed"):
        return True
    return bool(
        IDENTIFIER_EXPLICIT.match(stripped)
        or (IDENTIFIER_SUBJECT.search(stripped) and IDENTIFIER_SUFFIX.search(stripped))
    )


# Dates that are really "unset". These parse as text and fail as dates, which is
# why 11.5% of one real END_DATETIME column silently broke every duration.
SENTINEL_PATTERN = re.compile(
    r"^\s*(?:0{1,4}[-/]0{1,2}[-/]0{1,4}(?:[ t].*)?|0000-00-00.*|9999[-/].*|n/?a|null|nil|none|unknown|未知|无)\s*$",
    re.I,
)
# One cell carrying several values: 129/74, a pipe-delimited comorbidity list.
COMPOSITE_PATTERN = re.compile(r"^[^|;,/\\]+(?:\s*[|;/\\]\s*[^|;,/\\]+)+$")
INTEGER_PATTERN = re.compile(r"^[+-]?\d+$")
NUMBER_PATTERN = re.compile(r"^[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?$")
# Both orders, because a real extract exported day-first: matching only
# year-first left 915 of one column's timestamps looking like composite values,
# which buried the genuine composites — the pipe-delimited comorbidity strings —
# under an entire column of false positives.
DATE_PATTERN = re.compile(
    r"^(?:\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{4})"
    r"(?:[ tT]\d{1,2}:\d{2}(?::\d{2})?)?\s*$",
)


def is_blank(value: str) -> bool:
    return not value or not value.strip()


def infer_type(values: list[str]) -> str:
    """The narrowest type every non-blank value satisfies."""
    if not values:
        return "empty"
    if all(INTEGER_PATTERN.match(v.strip()) for v in values):
        return "integer"
    if all(NUMBER_PATTERN.match(v.strip()) for v in values):
        return "number"
    if all(DATE_PATTERN.match(v.strip()) for v in values):
        return "date"
    return "text"


def profile_column(name: str, cells: list[str]) -> tuple[dict, set[str]]:
    """Returns the emitted profile and, separately, the distinct values.

    The values are used to compute join reachability and are never written out.
    """
    filled = [c for c in cells if not is_blank(c)]
    stripped = [c.strip() for c in filled]
    counts = Counter(stripped)
    distinct = len(counts)
    identifying = is_identifying(name)
    complete = distinct <= VOCABULARY_COMPLETE_MAX and not identifying
    shown = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
    if identifying:
        shown = []
    elif not complete:
        shown = shown[:VOCABULARY_SAMPLE]
    sentinels = sorted({v for v in stripped if SENTINEL_PATTERN.match(v)})
    # A date is slash-separated and a sentinel date is both; neither is a cell
    # carrying two values, and reporting them as composite buries the real ones.
    composites = [
        v for v in stripped
        if COMPOSITE_PATTERN.match(v) and not SENTINEL_PATTERN.match(v) and not DATE_PATTERN.match(v)
    ]
    return {
        "name": name,
        "rows": len(cells),
        "filled": len(filled),
        # Density completeness in Weiskopf's sense: the proportion of rows where
        # this field carries a value. It says nothing about whether the value is
        # correct, and it is not the other three completeness definitions.
        "densityCompleteness": round(len(filled) / len(cells), 4) if cells else 0.0,
        "distinct": distinct,
        "inferredType": infer_type(stripped),
        "vocabulary": {"complete": complete, "identifying": identifying, "values": [[v, c] for v, c in shown]},
        "sentinelSuspects": sentinels,
        "compositeSuspects": {
            "count": len(composites),
            "examples": [] if identifying else sorted({v for v in composites})[:3],
        },
    }, set(counts)


def read_delimited(path: Path, delimiter: str) -> list[tuple[str, list[str], list[list[str]]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=delimiter))
    if not rows:
        return [(path.name, [], [])]
    return [(path.name, [str(h) for h in rows[0]], [[str(c) for c in r] for r in rows[1:]])]


def read_excel(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - environment dependent
        raise SystemExit(
            f"{path.name} is an Excel workbook and openpyxl is not installed. "
            "Install openpyxl, or export each sheet to CSV and profile those."
        )
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables = []
    for sheet in workbook.worksheets:
        rows = [["" if c is None else str(c) for c in row] for row in sheet.iter_rows(values_only=True)]
        if not rows:
            tables.append((f"{path.name}#{sheet.title}", [], []))
            continue
        header = [
            f"unnamed_{index + 1}" if not str(h).strip() else str(h)
            for index, h in enumerate(rows[0])
        ]
        tables.append((f"{path.name}#{sheet.title}", header, rows[1:]))
    workbook.close()
    return tables


def read_table(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel(path)
    if suffix in {".tsv", ".tab"}:
        return read_delimited(path, "\t")
    return read_delimited(path, ",")


def fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return f"sha256:{digest.hexdigest()}"


def profile_tables(paths: list[Path]) -> tuple[dict, dict[tuple[str, str], set[str]]]:
    tables = []
    values: dict[tuple[str, str], set[str]] = {}
    for path in paths:
        digest = fingerprint(path)
        for name, header, rows in read_table(path):
            columns = []
            for index, column_name in enumerate(header):
                cells = [row[index] if index < len(row) else "" for row in rows]
                profile, distinct = profile_column(str(column_name), cells)
                columns.append(profile)
                values[(name, str(column_name))] = distinct
            tables.append({
                "name": name,
                "sourceFile": path.name,
                "sourceFingerprint": digest,
                "rows": len(rows),
                "columns": columns,
            })
    return {"schemaVersion": SCHEMA_VERSION, "tables": tables}, values


def discover_joins(values: dict[tuple[str, str], set[str]]) -> list[dict]:
    """Inclusion dependencies: is every value of A.col also a value of B.col?

    Computed from the real values, which are never emitted. A key that is
    present but empty is the finding this exists to surface — one real diagnosis
    table carried PATIENT_ID at 0% fill, which no schema diagram would show.
    """
    candidates = {key: v for key, v in values.items() if JOIN_KEY_PATTERN.search(key[1])}
    joins = []
    for (left_table, column_name), left_values in sorted(candidates.items()):
        for (right_table, right_column), right_values in sorted(candidates.items()):
            if left_table >= right_table or column_name != right_column:
                continue
            if not left_values or not right_values:
                joins.append({
                    "left": f"{left_table}.{column_name}",
                    "right": f"{right_table}.{right_column}",
                    "reachable": False,
                    "reason": "one side has no values at all",
                    "containment": 0.0,
                })
                continue
            matched = len(left_values & right_values)
            joins.append({
                "left": f"{left_table}.{column_name}",
                "right": f"{right_table}.{right_column}",
                "reachable": matched > 0,
                "reason": "" if matched else "no value of the left key occurs in the right key",
                "containment": round(matched / len(left_values), 4),
            })
    return joins


def find_type_conflicts(tables: list[dict]) -> list[dict]:
    by_name: dict[str, dict[str, str]] = {}
    for table in tables:
        for column in table["columns"]:
            if column["inferredType"] == "empty":
                continue
            by_name.setdefault(column["name"], {})[table["name"]] = column["inferredType"]
    conflicts = []
    for column_name, types in sorted(by_name.items()):
        if len(set(types.values())) > 1:
            conflicts.append({"column": column_name, "types": dict(sorted(types.items()))})
    return conflicts


def render_markdown(profile: dict) -> str:
    lines = ["# 数据剖析（data profile）", ""]
    lines.append(
        "本文件由 `data-profile.py` 生成，数字全部可复算：对同一批输入重跑该脚本即得同一份 JSON。"
    )
    lines.append("")
    lines.append(
        "**填充率的口径**：本表的填充率是 Weiskopf 四义中的 **density completeness**"
        "（该字段在多少比例的行上有值），不涉及取值是否正确，也不是其余三种完整性。"
    )
    lines.append("")
    for table in profile["tables"]:
        lines.append(f"## {table['name']}")
        lines.append("")
        lines.append(f"- 来源文件：`{table['sourceFile']}`（`{table['sourceFingerprint']}`）")
        lines.append(f"- 行数：{table['rows']}；列数：{len(table['columns'])}")
        lines.append("")
        lines.append("| 字段 | 填充率 | 取值基数 | 推断类型 | 词表 | 备注 |")
        lines.append("|---|---|---|---|---|---|")
        for column in table["columns"]:
            vocabulary = column["vocabulary"]
            rendered = "、".join(f"`{v}`×{c}" for v, c in vocabulary["values"][:8])
            if vocabulary["identifying"]:
                rendered = "（标识符列，取值不外带）"
            elif not vocabulary["complete"]:
                rendered = f"（高基数，仅列高频）{rendered}"
            notes = []
            if column["sentinelSuspects"]:
                notes.append("哨兵值：" + "、".join(f"`{s}`" for s in column["sentinelSuspects"][:3]))
            if column["compositeSuspects"]["count"]:
                notes.append(f"复合值 {column['compositeSuspects']['count']} 处")
            if column["filled"] == 0:
                notes.append("**该列全空**")
            lines.append(
                f"| `{column['name']}` | {column['densityCompleteness']:.1%} "
                f"({column['filled']}/{column['rows']}) | {column['distinct']} | "
                f"{column['inferredType']} | {rendered or '—'} | {'；'.join(notes) or '—'} |"
            )
        lines.append("")

    lines.append("## 跨表连接可达性（inclusion dependency）")
    lines.append("")
    if not profile["joins"]:
        lines.append("未发现同名的候选连接键。")
    else:
        lines.append("| 左 | 右 | 可达 | 包含度 | 说明 |")
        lines.append("|---|---|---|---|---|")
        for join in profile["joins"]:
            reachable = {True: "✅", False: "❌", None: "⚠️"}[join["reachable"]]
            containment = "—" if join["containment"] is None else f"{join['containment']:.1%}"
            lines.append(
                f"| `{join['left']}` | `{join['right']}` | {reachable} | {containment} | {join['reason'] or '—'} |"
            )
    lines.append("")

    lines.append("## 跨表类型不一致")
    lines.append("")
    if not profile["typeConflicts"]:
        lines.append("未发现同名字段在不同表中被推断为不同类型。")
    else:
        lines.append("| 字段 | 各表推断类型 |")
        lines.append("|---|---|")
        for conflict in profile["typeConflicts"]:
            rendered = "；".join(f"{table}={kind}" for table, kind in conflict["types"].items())
            lines.append(f"| `{conflict['column']}` | {rendered} |")
    lines.append("")
    return "\n".join(lines)


# =====================================================================
# Phase 2 — domain-derived quantities for a therapeutic-drug-monitoring
# extract. These are the quantities that carry the information in this
# domain: dose-normalized concentration, position within the reference
# range, steady-state attainment, metabolite-to-parent ratio, and the
# longitudinal sampling structure. Everything here is computed from the
# workbook itself and is deterministic: the same input produces the same
# JSON, so the preflight's recomputation check covers it.
#
# Subjects are referenced only by pseudonyms assigned in a stable order:
# distinct CASE_NO sorted ascending -> P1, P2, ... The CASE_NO values
# themselves are never emitted.
# =====================================================================

import datetime as _dt


def _sheet_rows(paths: list[Path]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for path in paths:
        for name, header, rows in read_table(path):
            key = name.split("#", 1)[1] if "#" in name else name
            out[key] = [dict(zip(header, row)) for row in rows]
    return out


def _as_number(value) -> float | None:
    try:
        text = str(value).strip()
        if not text or text in {"无", "None", "null", "NaN"}:
            return None
        return float(text)
    except (ValueError, AttributeError):
        return None


def _parse_dayfirst(value) -> tuple[_dt.date, tuple[int, int] | None] | None:
    m = re.match(
        r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})(?:[ tT](\d{1,2}):(\d{2})(?::\d{2})?)?\s*$",
        str(value).strip(),
    )
    if not m:
        return None
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if day == 0 or month == 0:
        return None  # sentinel date
    try:
        date = _dt.date(year, month, day)
    except ValueError:
        return None
    time = None
    if m.group(4):
        time = (int(m.group(4)), int(m.group(5)))
    return (date, time)


def _parse_range(text) -> tuple[float, float] | None:
    m = re.match(r"^\s*([\d.]+)\s*-\s*([\d.]+)\s*$", str(text).strip())
    if not m:
        return None
    return (float(m.group(1)), float(m.group(2)))


def _range_status(value: float, rng) -> str | None:
    if rng is None:
        return None
    if value < rng[0]:
        return "below"
    if value > rng[1]:
        return "above"
    return "within"


# Frequency decoding. Plain codes decode directly. This extract also carries
# local duration-coded frequencies (QD11, QD12, QD16, BID4, BID5, BID8, TID4,
# BID1): a base frequency plus a numeric suffix that plausibly encodes the
# duration of the order in days. The base part determines the daily dose; the
# suffix does not. Neither reading is guaranteed by any field in the extract,
# so the script also computes a conservative "once-daily" scenario for the
# affected aripiprazole orders and reports dose and C/D as a range.
PLAIN_FREQ_MULT = {"QD": 1.0, "QN": 1.0, "QOD": 0.5, "BID": 2.0, "TID": 3.0,
                   "QID": 4.0, "ALWAYS": 1.0, "QHS": 1.0, "ONCE": None}
DURATION_CODED_BASE = {"QD": 1.0, "QN": 1.0, "BID": 2.0, "TID": 3.0}


def _freq_mult(freq) -> float | None:
    f = str(freq).strip().upper()
    if f in PLAIN_FREQ_MULT:
        return PLAIN_FREQ_MULT[f]
    for base, mult in DURATION_CODED_BASE.items():
        if f.startswith(base) and f[len(base):].isdigit():
            return mult
    return None


ARIPIPRAZOLE = "阿立哌唑"
ANALYTE_ORDER = ["阿立哌唑", "脱氢阿立哌唑", "总阿立哌唑", "奥氮平", "氯氮平",
                 "帕利哌酮（帕潘立酮）", "氯硝西泮"]


# =====================================================================
# Phase 6 planning quantities — minimum detectable effect and expected
# precision for the candidate estimands, computed from closed-form
# formulas with stated assumptions. Never a post-hoc power calculation:
# the dataset size is fixed, so the honest outputs are the MDE at a
# stated alpha and target power, and the expected interval width.
# =====================================================================

import math as _math


def compute_planning_quantities() -> dict:
    out = {"methodNote": "closed-form planning quantities; assumptions stated per item; "
                         "no post-hoc power is computed"}
    # (1) Riley anchor: continuous-outcome multivariable prediction model.
    # Published worked example (Riley et al., Stat Med 2019, PMID 30347470):
    # 25 predictor parameters -> at least 918 subjects (>=36.7 per parameter).
    riley_per_parameter = 36.7
    out["rileyContinuousOutcome"] = {
        "anchor": "Riley 2019 Part I worked example: 25 parameters -> >=918 subjects",
        "subjectsPerParameter": riley_per_parameter,
        "neededByPredictors": {
            str(p): _math.ceil(p * riley_per_parameter) for p in (3, 4, 6, 8, 10, 12)
        },
        "caveat": "the per-parameter ratio depends on the anticipated R^2; "
                  "lower R^2 raises the ratio. The anchor uses a modest R^2 scenario.",
    }
    # (2) Precision for the CYP2D6 effect on dose-normalized concentration.
    # Two independent groups (PM-like vs EM-like), outcome = ln(C/D),
    # assumed log-normal with CV. Difference of means; equal n per group.
    sigma = _math.sqrt(_math.log(1 + 0.4 ** 2))  # CV = 0.4 on the C/D scale
    z_alpha = 1.959964
    z_power = 0.841621
    table = []
    for n_per_group in (5, 10, 25, 50, 100, 150, 200):
        se = sigma * _math.sqrt(2.0 / n_per_group)
        mde_log = (z_alpha + z_power) * se          # alpha 0.05, power 0.80, two-sided
        ci_half_log = z_alpha * se                  # expected 95% CI half-width
        table.append({
            "nPerGroup": n_per_group,
            "mdeFoldRatio": round(_math.exp(mde_log), 2),
            "expected95CiHalfWidthOnLogScale": round(ci_half_log, 3),
            "expected95CiWidthInFoldRatio": round(_math.exp(2 * ci_half_log), 2),
        })
    out["mdeCiForCdRatioByCyp2D6"] = {
        "estimand": "ratio of geometric-mean dose-normalized concentration, PM-like vs EM-like",
        "assumptions": "two independent groups of equal size; ln(C/D) normal, CV=0.4 "
                       "(sigma=0.385); alpha=0.05 two-sided; target power 0.80 for the MDE",
        "table": table,
    }
    # (3) Expected precision of a single-sample dose-normalized concentration
    # audit: 95% prediction interval width for a future C/D observation given
    # a sample of n reference observations (t-interval with variance pooled
    # from the published population CV).
    pred = []
    for n_ref in (10, 25, 50, 100, 200, 500):
        half = z_alpha * sigma * _math.sqrt(1 + 1.0 / n_ref)
        pred.append({"nReferenceObservations": n_ref,
                     "95PiHalfWidthOnLogScale": round(half, 3),
                     "95PiWidthInFoldRatio": round(_math.exp(2 * half), 2)})
    out["predictionIntervalForCdAudit"] = {
        "estimand": "95% prediction interval width for an individual C/D observation "
                    "in a future patient, given n reference observations",
        "assumptions": "log-normal C/D with CV=0.4; interval computed on the log scale "
                       "with the population variance treated as known (z-based)",
        "table": pred,
    }
    # (4) External validation cohort for a continuous-outcome model:
    # Riley's binary-outcome external-validation criterion (>=100 events and
    # >=100 non-events, PMID 34031906) is cited in prose; the analogous
    # continuous-outcome requirement is a few hundred individuals. Stated as
    # an order-of-magnitude target rather than a formula.
    out["externalValidationScale"] = {
        "binaryOutcomeCriterion": ">=100 events and >=100 non-events (Riley 2021, PMID 34031906)",
        "continuousOutcomeOrderOfMagnitude": "several hundred individuals for stable calibration "
                                             "and small optimism; exact n depends on model complexity",
    }
    return out


def compute_domain_quantities(paths: list[Path]) -> dict:
    sheets = _sheet_rows(paths)
    for required in ("检验", "医嘱记录", "病案首页"):
        if required not in sheets:
            return {"applicable": False,
                    "reason": f"sheet {required} not present; domain quantities require a TDM extract"}
    labs = sheets["检验"]
    orders = sheets["医嘱记录"]
    front = sheets["病案首页"]
    vitals = sheets.get("体征", [])

    # --- pseudonym assignment: distinct CASE_NO ascending -> P1..Pn ------
    case_nos = sorted({str(r["CASE_NO"]) for r in labs if r.get("CASE_NO") is not None})
    pseudonym = {case: f"P{i + 1}" for i, case in enumerate(case_nos)}

    def pname(row) -> str | None:
        return pseudonym.get(str(row.get("CASE_NO")))

    # --- per-sample records ----------------------------------------------
    samples: dict[str, dict] = {}
    for row in labs:
        sn = str(row.get("SAMPLENO"))
        if sn not in samples:
            samples[sn] = {
                "sample": sn,
                "patient": pname(row),
                "applyDate": str(row.get("APPLY_DATE")),
                "testDate": str(row.get("TEST_DATE")),
                "analytes": [],
            }
        samples[sn]["analytes"].append({
            "analyte": str(row.get("PROJECT_NAME")),
            "value": _as_number(row.get("TEST_RESULT")),
            "range": _parse_range(str(row.get("REFFR_SCOPE"))) if row.get("REFFR_SCOPE") is not None else None,
        })
    sample_list = sorted(samples.values(), key=lambda s: (s["patient"], s["testDate"]))

    # --- aripiprazole dosing history --------------------------------------
    def is_ari(row) -> bool:
        content = str(row.get("ORDER_CONTENT") or "")
        name = str(row.get("DRUG_NAME") or "")
        return ARIPIPRAZOLE in content or ARIPIPRAZOLE in name

    ari_orders = []
    for row in orders:
        if str(row.get("DRUG_FLAG")) != "1" or not is_ari(row):
            continue
        way = str(row.get("MEDICATION_WAY") or "")
        if way in {"化验", "出院带药"}:
            continue
        start = _parse_dayfirst(row.get("START_DATETIME"))
        end = _parse_dayfirst(row.get("END_DATETIME"))
        if start is None:
            continue
        dose = _as_number(row.get("DOSAGE"))
        if dose is None:
            continue
        freq = str(row.get("FREQUENCY") or "")
        mult = _freq_mult(freq)
        ari_orders.append({
            "patient": pseudonym.get(str(row.get("CASE_NO"))),
            "doseMg": dose,
            "freq": freq,
            "multDecoded": mult,
            "start": start[0],
            "end": end[0] if end else None,
            "daysOnDose": None,
        })
    ari_orders.sort(key=lambda o: o["start"])

    def active_orders(patient, day):
        return [o for o in ari_orders
                if o["patient"] == patient and o["start"] <= day
                and (o["end"] is None or o["end"] >= day)]

    def dose_at(patient, day) -> dict:
        """Daily dose on the sampling day. lo/hi span the decoding ambiguity:
        decoded base frequency vs. conservative once-daily reading."""
        act = active_orders(patient, day)
        if not act:
            return {"known": False, "mgPerDayLo": None, "mgPerDayHi": None,
                    "orders": [], "undecodable": None}
        dose_hi, dose_lo = None, None
        undecodable = []
        for o in act:
            candidates = []
            if o["multDecoded"] is not None:
                candidates.append(o["doseMg"] * o["multDecoded"])
            if o["freq"].upper().startswith("BID") or o["freq"].upper() in {"QD11", "QD12", "QD16", "BID4", "BID5", "BID8", "TID4"}:
                candidates.append(o["doseMg"])  # conservative once-daily reading
            if not candidates:
                undecodable.append(o["freq"])
                continue
            dose_lo = min(candidates) if dose_lo is None else min(dose_lo, min(candidates))
            dose_hi = max(candidates) if dose_hi is None else max(dose_hi, max(candidates))
        return {"known": dose_hi is not None, "mgPerDayLo": dose_lo, "mgPerDayHi": dose_hi,
                "orders": [{"doseMg": o["doseMg"], "freq": o["freq"]} for o in act],
                "undecodable": undecodable or None}

    def first_ari_start(patient) -> _dt.date | None:
        starts = [o["start"] for o in ari_orders if o["patient"] == patient]
        return min(starts) if starts else None

    def days_since(d0, d1) -> int | None:
        return (d1 - d0).days if d0 is not None else None

    # --- steady-state and C/D per sample ----------------------------------
    per_sample = []
    for s in sample_list:
        parsed_test = _parse_dayfirst(s["testDate"])
        patient = s["patient"]
        dose = {"known": False, "mgPerDayLo": None, "mgPerDayHi": None}
        days_on_dose = None
        exposure_days = None
        if parsed_test:
            day = parsed_test[0]
            dose = dose_at(patient, day)
            act = active_orders(patient, day)
            if act:
                days_on_dose = min(days_since(o["start"], day) for o in act)
            first = first_ari_start(patient)
            exposure_days = days_since(first, day) if first else None
        steady = None
        if days_on_dose is not None:
            steady = "yes" if days_on_dose >= 15 else ("borderline" if days_on_dose >= 10 else "no")
        analyte_rows = []
        for a in s["analytes"]:
            analyte_rows.append({
                "analyte": a["analyte"],
                "value": a["value"],
                "range": a["range"],
                "status": _range_status(a["value"], a["range"]) if a["value"] is not None else None,
            })
        by_analyte = {a["analyte"]: a["value"] for a in s["analytes"]}
        parent = by_analyte.get("阿立哌唑")
        metabolite = by_analyte.get("脱氢阿立哌唑")
        total = by_analyte.get("总阿立哌唑")
        meta_parent = (round(metabolite / parent, 3)
                       if parent and metabolite and parent > 0 else None)
        cd = {}
        if dose["known"] and dose["mgPerDayLo"] and dose["mgPerDayHi"]:
            for label, conc in (("parent", parent), ("total", total)):
                if conc is None:
                    continue
                cd[label] = {
                    "perMgPerDayLo": round(conc / dose["mgPerDayHi"], 2),
                    "perMgPerDayHi": round(conc / dose["mgPerDayLo"], 2),
                }
        per_sample.append({
            "patient": patient,
            "testDate": s["testDate"],
            "applyDate": s["applyDate"],
            "analytes": analyte_rows,
            "ariDoseMgPerDayLo": dose["mgPerDayLo"],
            "ariDoseMgPerDayHi": dose["mgPerDayHi"],
            "doseKnown": dose["known"],
            "doseOrders": dose["orders"],
            "undecodableFrequencies": dose["undecodable"],
            "daysOnCurrentDose": days_on_dose,
            "daysAripiprazoleExposure": exposure_days,
            "steadyState": steady,
            "metaboliteParentRatio": meta_parent,
            "doseNormalized": cd,
        })

    # --- aggregates --------------------------------------------------------
    per_patient = {}
    for s in per_sample:
        p = s["patient"]
        per_patient.setdefault(p, {"samples": 0, "validDoseSamples": 0,
                                   "analytes": set(), "mainDiagnosis": None,
                                   "weightRecordsNearSample": []})
        per_patient[p]["samples"] += 1
        if s["doseKnown"]:
            per_patient[p]["validDoseSamples"] += 1
        for a in s["analytes"]:
            per_patient[p]["analytes"].add(a["analyte"])
    for row in front:
        p = pseudonym.get(str(row.get("CASE_NO")))
        if p in per_patient and row.get("MAIN_DIAGNOSIS_NAME"):
            per_patient[p]["mainDiagnosis"] = str(row.get("MAIN_DIAGNOSIS_NAME"))
    for row in vitals:
        p = pseudonym.get(str(row.get("CASE_NO")))
        if p not in per_patient or str(row.get("SIGN_TYPE")) != "体重":
            continue
        parsed = _parse_dayfirst(row.get("RECORD_DATE"))
        if not parsed:
            continue
        per_patient[p]["weightRecordsNearSample"].append(str(parsed[0]))

    cd_values_lo, cd_values_hi = [], []
    for s in per_sample:
        if "total" in s["doseNormalized"]:
            cd_values_lo.append(s["doseNormalized"]["total"]["perMgPerDayLo"])
            cd_values_hi.append(s["doseNormalized"]["total"]["perMgPerDayHi"])
    cd_spread = None
    if cd_values_lo and cd_values_hi:
        all_cd = cd_values_lo + cd_values_hi
        cd_spread = {
            "nSamples": len(cd_values_lo),
            "min": min(all_cd), "max": max(all_cd), "median": round(sorted(all_cd)[len(all_cd) // 2], 2),
            "maxFoldVsMin": round(max(all_cd) / min(all_cd), 2),
        }

    range_counts = {"within": 0, "below": 0, "above": 0, "noRange": 0}
    for s in per_sample:
        for a in s["analytes"]:
            if a["status"] is None:
                range_counts["noRange"] += 1
            else:
                range_counts[a["status"]] += 1

    return {
        "applicable": True,
        "pseudonymRule": "distinct CASE_NO sorted ascending -> P1..Pn; source values never emitted",
        "nPatients": len(per_patient),
        "nSamples": len(per_sample),
        "nSamplesPerPatient": {p: v["samples"] for p, v in sorted(per_patient.items())},
        "patientsWith2PlusSamples": sum(1 for v in per_patient.values() if v["samples"] >= 2),
        "patientsWith3PlusSamples": sum(1 for v in per_patient.values() if v["samples"] >= 3),
        "samplingClockTimePresent": any(_parse_dayfirst(s["testDate"])[1] is not None for s in per_sample),
        "perSample": per_sample,
        "perPatient": {
            p: {
                "samples": v["samples"],
                "validDoseSamples": v["validDoseSamples"],
                "analytes": sorted(v["analytes"]),
                "mainDiagnosis": v["mainDiagnosis"],
                "nWeightRecords": len(v["weightRecordsNearSample"]),
            }
            for p, v in sorted(per_patient.items())
        },
        "aripiprazoleAdminOrders": len(ari_orders),
        "aripiprazoleDoseCDFoldSpread": cd_spread,
        "analyteRangeStatusCounts": range_counts,
        "frequencyDecoding": {
            "plain": sorted(PLAIN_FREQ_MULT),
            "durationCoded": sorted(DURATION_CODED_BASE) + ["BID4", "QD11", "QD12", "QD16", "BID5", "BID8", "TID4", "BID1"],
            "note": "numeric suffix plausibly encodes order duration in days; base determines daily dose; "
                    "undecodable codes force a scenario range in dose and C/D",
        },
        "missingForModeling": [
            "sampling clock time (TEST_DATE carries date only)",
            "administration record (orders only; no administered/observed dose)",
            "adherence record",
            "outcome/rating-scale field",
            "CYP2D6/CYP3A4 genotype",
            "steady-state guarantee at sampling",
        ],
    }


def render_domain_markdown(dq: dict) -> list[str]:
    if not dq.get("applicable"):
        return ["## 领域量（domain quantities）", "",
                f"不适用：{dq.get('reason')}", ""]
    lines = ["## 领域量（Phase 2，TDM）", "",
             f"- 患者数（按 CASE_NO 去重，化名 P1–Pn）：{dq['nPatients']}；TDM 血样数：{dq['nSamples']}",
             f"- 每患者采样次数：{dq['nSamplesPerPatient']}",
             f"- 采样≥2 次的患者数：{dq['patientsWith2PlusSamples']}；采样≥3 次的患者数：{dq['patientsWith3PlusSamples']}",
             f"- 检验日期是否含时刻：{'是' if dq['samplingClockTimePresent'] else '否（仅有日期，无法判定谷浓度/峰浓度）'}",
             f"- 阿立哌唑给药医嘱数（剔除化验/出院带药）：{dq['aripiprazoleAdminOrders']}", ""]
    lines.append("| 化名 | 采血日 | 药物 | 浓度 | 参考区间 | 区间位置 | 日剂量(mg, 解码区间) | 稳态判定 | 代谢物/母药比 |")
    lines.append("|---|---|---|---|---|---|---|---|---|")
    for s in dq["perSample"]:
        for a in s["analytes"]:
            rng = f"{a['range'][0]:g}-{a['range'][1]:g}" if a["range"] else "—"
            dose = (f"{s['ariDoseMgPerDayLo']:g}-{s['ariDoseMgPerDayHi']:g}"
                    if s["doseKnown"] else "无法确定")
            lines.append(
                f"| {s['patient']} | {s['testDate']} | {a['analyte']} | {a['value']} | {rng} | "
                f"{a['status'] or '—'} | {dose} | {s['steadyState'] or '—'} | "
                f"{s['metaboliteParentRatio'] if a['analyte'] == '阿立哌唑' else ''} |"
            )
    lines.append("")
    cd = dq["aripiprazoleDoseCDFoldSpread"]
    if cd:
        lines.append(
            f"- **剂量校正浓度（总阿立哌唑，ng/mL per mg/day）**：n={cd['nSamples']}，"
            f"min={cd['min']:g}，max={cd['max']:g}，中位数={cd['median']:g}，最大/最小={cd['maxFoldVsMin']:g} 倍"
        )
    lines.append(f"- 区间位置计数（全部浓度测定）：{dq['analyteRangeStatusCounts']}")
    lines.append("- 频率解码：明文编码 " + "、".join(dq["frequencyDecoding"]["plain"]) +
                 "；时长编码 " + "、".join(dq["frequencyDecoding"]["durationCoded"]) +
                 "；无法解码的编码使日剂量与 C/D 只能以区间表示。")
    lines.append("- 建模所缺字段（据本剖析直接可判）：" + "、".join(dq["missingForModeling"]) + "。")
    lines.append("")
    return lines


def main() -> int:
    parser = argparse.ArgumentParser(description="Profile a dataset for research scoping.")
    parser.add_argument("inputs", nargs="+", help="CSV, TSV, or XLSX files to profile")
    parser.add_argument("--json", dest="json_out", default="data-profile.json")
    parser.add_argument("--markdown", dest="markdown_out", default="data-profile.md")
    args = parser.parse_args()

    paths = [Path(p) for p in args.inputs]
    missing = [str(p) for p in paths if not p.is_file()]
    if missing:
        raise SystemExit("input file(s) not found: " + ", ".join(missing))

    profile, values = profile_tables(sorted(paths, key=lambda p: p.name))
    profile["joins"] = discover_joins(values)
    profile["typeConflicts"] = find_type_conflicts(profile["tables"])
    profile["domainQuantities"] = compute_domain_quantities(sorted(paths, key=lambda p: p.name))
    profile["planningQuantities"] = compute_planning_quantities()

    Path(args.json_out).write_text(json.dumps(profile, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    Path(args.markdown_out).write_text(render_markdown(profile) + "\n" + "\n".join(render_domain_markdown(profile["domainQuantities"])), encoding="utf-8")
    empty_columns = sum(1 for t in profile["tables"] for c in t["columns"] if c["filled"] == 0)
    print(
        f"profiled {len(profile['tables'])} table(s), "
        f"{sum(len(t['columns']) for t in profile['tables'])} column(s), "
        f"{empty_columns} entirely empty, "
        f"{len(profile['joins'])} candidate join(s), "
        f"{len(profile['typeConflicts'])} type conflict(s)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())

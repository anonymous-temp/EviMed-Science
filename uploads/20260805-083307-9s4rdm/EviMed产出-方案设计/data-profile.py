#!/usr/bin/env python3
"""Mechanical data profiling for dataset research scoping.

Single-column profiling, inclusion-dependency discovery across tables, and the
trap detectors that a real hospital extract needed (sentinel dates, a join key
that is present but empty, a column typed differently in two tables, composite
values in one cell).

Deterministic by construction: the same inputs produce byte-identical JSON, so
the preflight can re-run this and compare. Nothing here reads the network, and
no cell value is printed except as an aggregate or a vocabulary entry.

Copy this file into the workspace as data-profile.py and run it there, so the
deliverable regenerates from a deliverable:

    python3 data-profile.py data/*.xlsx --json data-profile.json --markdown data-profile.md
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path

SCHEMA_VERSION = 1
# Above this many distinct values a vocabulary is reported as its most frequent
# entries rather than in full; below it, the full vocabulary is the finding.
VOCABULARY_COMPLETE_MAX = 30
VOCABULARY_SAMPLE = 15
# A column whose name looks like a join key is checked for containment in every
# other table that has a column of the same name.
JOIN_KEY_PATTERN = re.compile(r"(?:^|_)(?:id|no|code|key|num|number)$", re.I)
# Columns whose values identify a person or an episode of care. Their cardinality
# and fill rate are findings; their values are not. Printing the vocabulary of a
# five-patient PATIENT_ID column writes five real hospital numbers into the
# deliverable — which is what the preflight's first gate exists to catch, and it
# caught this script doing it. Joins are still computed from the real values;
# they are simply never emitted.
#
# Matching the subject word alone was too broad: RECORD_DATE and RECORD_CONTENT
# contain "record", so a vital-signs table had its most analytic column masked.
# An identifier is a subject word carrying an id-shaped suffix, or one of the
# few names that are identifiers outright.
IDENTIFIER_EXPLICIT = re.compile(r"^(?:id|mrn|姓名|患者姓名|身份证号?|病案号|住院号|门诊号|就诊号|med_rec_no|sampleno)$", re.I)
IDENTIFIER_SUBJECT = re.compile(
    r"(?:patient|subject|person|case|record|admission|visit|encounter|inpatient|"
    r"病案|住院|门诊|患者|病人|就诊)",
    re.I,
)
IDENTIFIER_SUFFIX = re.compile(r"(?:^|_)(?:id|no|num|number|code)$", re.I)


def is_identifying(name: str) -> bool:
    """True when a column's values identify a person or an episode of care."""
    return bool(
        IDENTIFIER_EXPLICIT.match(name.strip())
        or (IDENTIFIER_SUBJECT.search(name) and IDENTIFIER_SUFFIX.search(name))
    )


# Dates that are really "unset". These parse as text and fail as dates, which is
# why 11.5% of one real END_DATETIME column silently broke every duration.
SENTINEL_PATTERN = re.compile(
    r"^\s*(?:0{1,4}[-/]0{1,2}[-/]0{1,4}(?:[ t].*)?|0000-00-00.*|9999[-/].*|n/?a|null|nil|none|unknown|未知|无)\s*$",
    re.I,
)
# One cell carrying several values: 129/74, a pipe-delimited comorbidity list.
COMPOSITE_PATTERN = re.compile(r"^[^|;,/\\]+(?:\s*[|;/\\]\s*[^|;,/\\]+)+$")
INTEGER_PATTERN = re.compile(r"^[+-]?\d+$")
NUMBER_PATTERN = re.compile(r"^[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?$")
# Both orders, because a real extract exported day-first: matching only
# year-first left 915 of one column's timestamps looking like composite values,
# which buried the genuine composites — the pipe-delimited comorbidity strings —
# under an entire column of false positives.
DATE_PATTERN = re.compile(
    r"^(?:\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{4})"
    r"(?:[ tT]\d{1,2}:\d{2}(?::\d{2})?)?\s*$",
)


def is_blank(value: str) -> bool:
    return not value or not value.strip()


def infer_type(values: list[str]) -> str:
    """The narrowest type every non-blank value satisfies."""
    if not values:
        return "empty"
    if all(INTEGER_PATTERN.match(v.strip()) for v in values):
        return "integer"
    if all(NUMBER_PATTERN.match(v.strip()) for v in values):
        return "number"
    if all(DATE_PATTERN.match(v.strip()) for v in values):
        return "date"
    return "text"


def profile_column(name: str, cells: list[str]) -> tuple[dict, set[str]]:
    """Returns the emitted profile and, separately, the distinct values.

    The values are used to compute join reachability and are never written out.
    """
    filled = [c for c in cells if not is_blank(c)]
    stripped = [c.strip() for c in filled]
    counts = Counter(stripped)
    distinct = len(counts)
    identifying = is_identifying(name)
    complete = distinct <= VOCABULARY_COMPLETE_MAX and not identifying
    shown = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
    if identifying:
        shown = []
    elif not complete:
        shown = shown[:VOCABULARY_SAMPLE]
    sentinels = sorted({v for v in stripped if SENTINEL_PATTERN.match(v)})
    # A date is slash-separated and a sentinel date is both; neither is a cell
    # carrying two values, and reporting them as composite buries the real ones.
    composites = [
        v for v in stripped
        if COMPOSITE_PATTERN.match(v) and not SENTINEL_PATTERN.match(v) and not DATE_PATTERN.match(v)
    ]
    return {
        "name": name,
        "rows": len(cells),
        "filled": len(filled),
        # Density completeness in Weiskopf's sense: the proportion of rows where
        # this field carries a value. It says nothing about whether the value is
        # correct, and it is not the other three completeness definitions.
        "densityCompleteness": round(len(filled) / len(cells), 4) if cells else 0.0,
        "distinct": distinct,
        "inferredType": infer_type(stripped),
        "vocabulary": {"complete": complete, "identifying": identifying, "values": [[v, c] for v, c in shown]},
        "sentinelSuspects": sentinels,
        "compositeSuspects": {
            "count": len(composites),
            "examples": [] if identifying else sorted({v for v in composites})[:3],
        },
    }, set(counts)


def read_delimited(path: Path, delimiter: str) -> list[tuple[str, list[str], list[list[str]]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=delimiter))
    if not rows:
        return [(path.name, [], [])]
    return [(path.name, [str(h) for h in rows[0]], [[str(c) for c in r] for r in rows[1:]])]


def read_excel(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - environment dependent
        raise SystemExit(
            f"{path.name} is an Excel workbook and openpyxl is not installed. "
            "Install openpyxl, or export each sheet to CSV and profile those."
        )
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables = []
    for sheet in workbook.worksheets:
        rows = [["" if c is None else str(c) for c in row] for row in sheet.iter_rows(values_only=True)]
        if not rows:
            tables.append((f"{path.name}#{sheet.title}", [], []))
            continue
        tables.append((f"{path.name}#{sheet.title}", rows[0], rows[1:]))
    workbook.close()
    return tables


def read_table(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel(path)
    if suffix in {".tsv", ".tab"}:
        return read_delimited(path, "\t")
    return read_delimited(path, ",")


def fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return f"sha256:{digest.hexdigest()}"


def profile_tables(paths: list[Path]) -> tuple[dict, dict[tuple[str, str], set[str]]]:
    tables = []
    values: dict[tuple[str, str], set[str]] = {}
    for path in paths:
        digest = fingerprint(path)
        for name, header, rows in read_table(path):
            columns = []
            for index, column_name in enumerate(header):
                cells = [row[index] if index < len(row) else "" for row in rows]
                profile, distinct = profile_column(str(column_name), cells)
                columns.append(profile)
                values[(name, str(column_name))] = distinct
            tables.append({
                "name": name,
                "sourceFile": path.name,
                "sourceFingerprint": digest,
                "rows": len(rows),
                "columns": columns,
            })
    return {"schemaVersion": SCHEMA_VERSION, "tables": tables}, values


def discover_joins(values: dict[tuple[str, str], set[str]]) -> list[dict]:
    """Inclusion dependencies: is every value of A.col also a value of B.col?

    Computed from the real values, which are never emitted. A key that is
    present but empty is the finding this exists to surface — one real diagnosis
    table carried PATIENT_ID at 0% fill, which no schema diagram would show.
    """
    candidates = {key: v for key, v in values.items() if JOIN_KEY_PATTERN.search(key[1])}
    joins = []
    for (left_table, column_name), left_values in sorted(candidates.items()):
        for (right_table, right_column), right_values in sorted(candidates.items()):
            if left_table >= right_table or column_name != right_column:
                continue
            if not left_values or not right_values:
                joins.append({
                    "left": f"{left_table}.{column_name}",
                    "right": f"{right_table}.{right_column}",
                    "reachable": False,
                    "reason": "one side has no values at all",
                    "containment": 0.0,
                })
                continue
            matched = len(left_values & right_values)
            joins.append({
                "left": f"{left_table}.{column_name}",
                "right": f"{right_table}.{right_column}",
                "reachable": matched > 0,
                "reason": "" if matched else "no value of the left key occurs in the right key",
                "containment": round(matched / len(left_values), 4),
            })
    return joins


def mask_subset_identifier_columns(tables: list[dict], values: dict[tuple[str, str], set[str]]) -> None:
    """Mask columns whose distinct values are entirely contained in an
    identifying column's values — they carry the same identifiers under a
    different name. A real hospital extract stored CASE_NO values in an unnamed
    vital-signs column; unmasked, the profile's vocabulary printed five real
    case numbers that the preflight then correctly flagged as leaked."""
    identifying_sets: list[set[str]] = []
    for table in tables:
        for column in table["columns"]:
            if column["vocabulary"]["identifying"]:
                key = (table["name"], column["name"])
                if key in values:
                    identifying_sets.append(values[key])
    if not identifying_sets:
        return
    for table in tables:
        for column in table["columns"]:
            if column["vocabulary"]["identifying"]:
                continue
            key = (table["name"], column["name"])
            if key not in values or not values[key]:
                continue
            if any(values[key] <= ident for ident in identifying_sets):
                column["vocabulary"]["identifying"] = True
                column["vocabulary"]["values"] = []
                column["vocabulary"]["complete"] = True
                column["note"] = "值完全包含于某标识符列取值 → 按标识符掩码"
                column["compositeSuspects"] = {"count": 0, "examples": []}
                column["sentinelSuspects"] = []


def find_type_conflicts(tables: list[dict]) -> list[dict]:
    by_name: dict[str, dict[str, str]] = {}
    for table in tables:
        for column in table["columns"]:
            if column["inferredType"] == "empty":
                continue
            by_name.setdefault(column["name"], {})[table["name"]] = column["inferredType"]
    conflicts = []
    for column_name, types in sorted(by_name.items()):
        if len(set(types.values())) > 1:
            conflicts.append({"column": column_name, "types": dict(sorted(types.items()))})
    return conflicts


def render_markdown(profile: dict) -> str:
    lines = ["# 数据剖析（data profile）", ""]
    lines.append(
        "本文件由 `data-profile.py` 生成，数字全部可复算：对同一批输入重跑该脚本即得同一份 JSON。"
    )
    lines.append("")
    lines.append(
        "**填充率的口径**：本表的填充率是 Weiskopf 四义中的 **density completeness**"
        "（该字段在多少比例的行上有值），不涉及取值是否正确，也不是其余三种完整性。"
    )
    lines.append("")
    for table in profile["tables"]:
        lines.append(f"## {table['name']}")
        lines.append("")
        lines.append(f"- 来源文件：`{table['sourceFile']}`（`{table['sourceFingerprint']}`）")
        lines.append(f"- 行数：{table['rows']}；列数：{len(table['columns'])}")
        lines.append("")
        lines.append("| 字段 | 填充率 | 取值基数 | 推断类型 | 词表 | 备注 |")
        lines.append("|---|---|---|---|---|---|")
        for column in table["columns"]:
            vocabulary = column["vocabulary"]
            rendered = "、".join(f"`{v}`×{c}" for v, c in vocabulary["values"][:8])
            if vocabulary["identifying"]:
                rendered = "（标识符列，取值不外带）"
            elif not vocabulary["complete"]:
                rendered = f"（高基数，仅列高频）{rendered}"
            notes = []
            if column["sentinelSuspects"]:
                notes.append("哨兵值：" + "、".join(f"`{s}`" for s in column["sentinelSuspects"][:3]))
            if column["compositeSuspects"]["count"]:
                notes.append(f"复合值 {column['compositeSuspects']['count']} 处")
            if column["filled"] == 0:
                notes.append("**该列全空**")
            lines.append(
                f"| `{column['name']}` | {column['densityCompleteness']:.1%} "
                f"({column['filled']}/{column['rows']}) | {column['distinct']} | "
                f"{column['inferredType']} | {rendered or '—'} | {'；'.join(notes) or '—'} |"
            )
        lines.append("")

    lines.append("## 跨表连接可达性（inclusion dependency）")
    lines.append("")
    if not profile["joins"]:
        lines.append("未发现同名的候选连接键。")
    else:
        lines.append("| 左 | 右 | 可达 | 包含度 | 说明 |")
        lines.append("|---|---|---|---|---|")
        for join in profile["joins"]:
            reachable = {True: "✅", False: "❌", None: "⚠️"}[join["reachable"]]
            containment = "—" if join["containment"] is None else f"{join['containment']:.1%}"
            lines.append(
                f"| `{join['left']}` | `{join['right']}` | {reachable} | {containment} | {join['reason'] or '—'} |"
            )
    lines.append("")

    lines.append("## 跨表类型不一致")
    lines.append("")
    if not profile["typeConflicts"]:
        lines.append("未发现同名字段在不同表中被推断为不同类型。")
    else:
        lines.append("| 字段 | 各表推断类型 |")
        lines.append("|---|---|")
        for conflict in profile["typeConflicts"]:
            rendered = "；".join(f"{table}={kind}" for table, kind in conflict["types"].items())
            lines.append(f"| `{conflict['column']}` | {rendered} |")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Profile a dataset for research scoping.")
    parser.add_argument("inputs", nargs="+", help="CSV, TSV, or XLSX files to profile")
    parser.add_argument("--json", dest="json_out", default="data-profile.json")
    parser.add_argument("--markdown", dest="markdown_out", default="data-profile.md")
    args = parser.parse_args()

    paths = [Path(p) for p in args.inputs]
    missing = [str(p) for p in paths if not p.is_file()]
    if missing:
        raise SystemExit("input file(s) not found: " + ", ".join(missing))

    profile, values = profile_tables(sorted(paths, key=lambda p: p.name))
    profile["joins"] = discover_joins(values)
    profile["typeConflicts"] = find_type_conflicts(profile["tables"])
    mask_subset_identifier_columns(profile["tables"], values)

    Path(args.json_out).write_text(json.dumps(profile, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    Path(args.markdown_out).write_text(render_markdown(profile), encoding="utf-8")
    empty_columns = sum(1 for t in profile["tables"] for c in t["columns"] if c["filled"] == 0)
    print(
        f"profiled {len(profile['tables'])} table(s), "
        f"{sum(len(t['columns']) for t in profile['tables'])} column(s), "
        f"{empty_columns} entirely empty, "
        f"{len(profile['joins'])} candidate join(s), "
        f"{len(profile['typeConflicts'])} type conflict(s)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())

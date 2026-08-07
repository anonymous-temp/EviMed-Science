#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tdm_dose_timeline.py — 重建每患者目标药物剂量时间线,计算采样日剂量、C/D 比、
代谢物/母药比、体重标准化剂量。输出: tdm_signal_result.json
"""
import json
import re
from collections import OrderedDict, defaultdict
from datetime import datetime, timedelta

import openpyxl

SRC = "/workspace/20260803TDM.xlsx"
OUT = "/workspace/tdm_signal_result.json"

FREQ_DAILY = {  # 本地编码 -> 每日次数 (主分析用;敏感性分析在报告里做)
    "QD": 1, "QN": 1, "BID": 2, "TID": 3, "QID": 4,
    "QD11": 1, "QD12": 1, "QD16": 1, "QOD": 0.5,
    "BID1": 2, "BID4": 2, "BID5": 2, "BID8": 2, "TID4": 3,
}
# 代码后缀解析: "QD12"=每天12点给药,"BID8"=每天8点给药(一次),疑似时间点编码
# BIDx/TIDx 保守按 BID/TID 计;若为"间隔小时数",BID8→3次/日(8h间隔),TID4→6次/日(4h间隔)
FREQ_DAILY_ALT = {  # 备选映射(间隔小时解读): BID8=3, TID4=6
    "QD11": 1, "QD12": 1, "QD16": 1, "BID1": 2, "BID4": 2, "BID5": 2,
    "BID8": 3, "TID4": 6,
}

TARGET = ["阿立哌唑", "氯氮平", "奥氮平", "帕利哌酮", "氯硝西泮", "碳酸锂"]
LAB_REF = {  # 检验科参考范围 (ng/mL) + AGNP/文献参考范围
    "阿立哌唑": (100, 350),
    "脱氢阿立哌唑": (None, None),
    "总阿立哌唑": (150, 500),
    "奥氮平": (20, 80),
    "氯氮平": (350, 600),
    "帕利哌酮": (20, 60),
    "氯硝西泮": (4, 80),
}


def parse_dt(s, date_only=False):
    s = str(s).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    try:
        return datetime.strptime(s, "%d/%m/%Y")
    except ValueError:
        return None


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    adm = list(wb["病案首页"].iter_rows(values_only=True))
    ah = [str(h) for h in adm[0]]
    adm_rows = [dict(zip(ah, r)) for r in adm[1:]]
    ords = list(wb["医嘱记录"].iter_rows(values_only=True))
    oh = [str(h) for h in ords[0]]
    ord_rows = [dict(zip(oh, r)) for r in ords[1:]]
    labs = list(wb["检验"].iter_rows(values_only=True))
    lh = [str(h) for h in labs[0]]
    lab_rows = [dict(zip(lh, r)) for r in labs[1:]]
    vit = list(wb["体征"].iter_rows(values_only=True))
    vh = [str(h) for h in vit[0]]
    vit_rows = [dict(zip(vh, r)) for r in vit[1:]]
    wb.close()

    r = {"samples": [], "per_patient_summary": {}, "notes": {}}

    # 1) 体重: 每患者最后一次体重(kg)
    weight = {}
    for v in vit_rows:
        if str(v["SIGN_TYPE"]) == "体重":
            pid = str(v["PATIENT_ID"]).strip()
            dt = parse_dt(v["RECORD_DATE"])
            val = v["RECORD_CONTENT"]
            try:
                val = float(val)
            except (TypeError, ValueError):
                continue
            if pid not in weight or (dt and dt > weight[pid][0]):
                weight[pid] = (dt, val)
    r["notes"]["weight_kg"] = {k: v[1] for k, v in weight.items()}

    # 2) 目标药物医嘱(排除出院带药/ONCE瞬时单)
    drug_orders = []
    for o in ord_rows:
        if o.get("DRUG_FLAG") != 1:
            continue
        name = str(o["DRUG_NAME"])
        if not any(t in name for t in TARGET):
            continue
        way = str(o["MEDICATION_WAY"])
        freq = str(o["FREQUENCY"])
        if "出院带药" in way or freq == "ONCE":
            continue
        drug_orders.append({
            "pid": str(o["PATIENT_ID"]).strip(), "name": name, "dosage": float(o["DOSAGE"]),
            "unit": str(o["DOSAGE_UNIT"]), "freq": freq,
            "start": parse_dt(o["START_DATETIME"]),
            "end": parse_dt(o["END_DATETIME"]),
        })
    # 修正哨兵结束时间 0/0/0
    for d in drug_orders:
        if d["end"] and d["end"].year < 2000:
            d["end"] = None

    # 3) 采样批次
    by_sample = defaultdict(list)
    sample_info = {}
    for l in lab_rows:
        sn = str(l["SAMPLENO"])
        sample_info.setdefault(sn, {
            "pid": str(l["PATIENT_ID"]).strip(),
            "test_date": parse_dt(l["TEST_DATE"], date_only=True),
            "apply_date": parse_dt(l["APPLY_DATE"], date_only=True),
            "purpose": str(l["TEST_PURPOSE"]),
        })
        by_sample[sn].append({"project": str(l["PROJECT_NAME"]),
                              "result": float(l["TEST_RESULT"]), "ref": l["REFFR_SCOPE"]})

    # 4) 采样日剂量(在院期间覆盖采样日的口服医嘱之和)
    def daily_dose_at(pid, drug_key, date, freq_map):
        dd = 0.0
        for d in drug_orders:
            if d["pid"] != pid or drug_key not in d["name"]:
                continue
            if d["start"] and d["start"].date() <= date and (d["end"] is None or d["end"].date() >= date):
                n = freq_map.get(d["freq"])
                if n is None:
                    return None  # 无法映射 -> 剂量不可得
                dd += d["dosage"] * n
        return dd

    # 5) 逐采样批次: 浓度、剂量、C/D、活性部分
    samples_out = []
    for sn in sorted(sample_info):
        info = sample_info[sn]
        pid, tdate = info["pid"], info["test_date"]
        results = {x["project"]: x for x in by_sample[sn]}
        entry = {"sample_no": sn, "pid": pid, "test_date": str(tdate.date()),
                 "apply_date": str(info["apply_date"].date()), "results": {},
                 "dose_mg_day_main": {}, "dose_mg_day_alt": {}, "c_d_ratio": {}, "cd_weight": {}}
        for proj, x in results.items():
            entry["results"][proj] = {"value": x["result"], "ref": x["ref"]}
        # 剂量
        for key, label in [("阿立哌唑", "ari"), ("氯氮平", "clo"), ("奥氮平", "ola"),
                           ("帕利哌酮", "pal"), ("氯硝西泮", "clon"), ("碳酸锂", "lith")]:
            d_main = daily_dose_at(pid, key, tdate.date(), FREQ_DAILY)
            d_alt = daily_dose_at(pid, key, tdate.date(), FREQ_DAILY_ALT)
            entry["dose_mg_day_main"][label] = d_main
            entry["dose_mg_day_alt"][label] = d_alt
        # C/D: 母药浓度/日剂量 (ng/mL per mg/day)
        for label, proj in [("ari", "阿立哌唑"), ("clo", "氯氮平"), ("ola", "奥氮平"),
                            ("pal", "帕利哌酮（帕潘立酮）"), ("clon", "氯硝西泮")]:
            c = entry["results"].get(proj)
            d = entry["dose_mg_day_main"].get(label)
            if c and d:
                entry["c_d_ratio"][label] = round(c["value"] / d, 2)
        # 代谢物/母药比
        ari = entry["results"].get("阿立哌唑")
        dh = entry["results"].get("脱氢阿立哌唑")
        tot = entry["results"].get("总阿立哌唑")
        if ari and dh:
            entry["dhc_ari_ratio"] = round(dh["value"] / ari["value"], 3)
            entry["active_total"] = round(ari["value"] + dh["value"], 1)
            entry["total_measured"] = tot["value"] if tot else None
        # 体重标准化
        w = weight.get(pid, (None, None))[1]
        if w and entry["dose_mg_day_main"].get("ari"):
            entry["cd_weight"]["weight_kg"] = w
            entry["cd_weight"]["ari_dose_per_kg"] = round(entry["dose_mg_day_main"]["ari"] / w, 3)
        samples_out.append(entry)
    r["samples"] = samples_out

    # 6) 每患者汇总
    for a in adm_rows:
        pid = str(a["PATIENT_ID"]).strip()
        r["per_patient_summary"][pid] = {
            "age": a["AGE"], "sex": "女" if a["SEX"] == 2 else "男",
            "main_dx": a["MAIN_DIAGNOSIS_NAME"],
            "other_dx": a["OTHER_DIAGNOSIS_NAME"],
            "n_samples": sum(1 for s in samples_out if s["pid"] == pid),
        }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(r, f, ensure_ascii=False, indent=1, default=str)
    for s in samples_out:
        print(s["sample_no"], s["pid"], s["test_date"],
              "| dose:", s["dose_mg_day_main"],
              "| CD:", s.get("c_d_ratio"),
              "| DHC/ARI:", s.get("dhc_ari_ratio"), "| active:", s.get("active_total"),
              "| tot_measured:", s.get("total_measured"),
              "| wgt:", s.get("cd_weight"))
    print("\nweight:", r["notes"]["weight_kg"])


if __name__ == "__main__":
    main()

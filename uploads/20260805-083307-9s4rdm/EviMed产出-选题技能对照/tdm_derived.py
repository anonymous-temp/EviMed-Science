#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tdm_derived.py — TDM 派生量计算与信号可分析性评估
派生量: 每日剂量(频次映射后)、浓度/剂量比(C/D)、活性部分浓度、治疗范围达标、
采样间隔、合并用药数、体质指数不可得性检查、血压解析。
"""
import json
import re
from collections import OrderedDict, Counter, defaultdict

import openpyxl

SRC = "/workspace/20260803TDM.xlsx"
OUT = "/workspace/tdm_derived_result.json"

# 频次本地编码 -> 每日次数(保守/上限两种映射,供敏感性分析)
FREQ_DAILY_LO = {
    "QD": 1, "QN": 1, "BID": 2, "TID": 3, "QID": 4, "Q6H": 4, "Q8H": 3, "Q12H": 2,
    "QOD": 0.5, "BIW": 2 / 7, "TIW": 3 / 7, "QW": 1 / 7, "QM": 1 / 30,
    "ONCE": None, "ALWAYS": None, "ST": None, "PRN": None,
}
# 阿立哌唑/奥氮平/氯氮平/帕利哌酮 治疗参考范围 (ng/mL) — AGNP 共识/文献常用
THERAPEUTIC_RANGE = {
    "阿立哌唑": (100, 350),
    "脱氢阿立哌唑": (40, 140),  # 活性代谢物,参考范围有争议;记录值 20-80 为检验科范围
    "总阿立哌唑": (200, 700),   # 母药+代谢物;检验科标 150-500
    "奥氮平": (20, 80),
    "氯氮平": (350, 600),
    "帕利哌酮": (20, 60),
}

TDM_DRUGS = {"阿立哌唑", "脱氢阿立哌唑", "总阿立哌唑", "奥氮平", "氯氮平", "帕利哌酮", "氯硝西泮"}


def parse_dt(s):
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return __import__("datetime").datetime.strptime(s, fmt)
        except ValueError:
            pass
    try:
        return __import__("datetime").datetime.strptime(s, "%d/%m/%Y")
    except ValueError:
        return None


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    adm = list(wb["病案首页"].iter_rows(values_only=True))
    adm_header = [str(h) for h in adm[0]]
    adm_rows = [dict(zip(adm_header, r)) for r in adm[1:]]

    ords = list(wb["医嘱记录"].iter_rows(values_only=True))
    ord_header = [str(h) for h in ords[0]]
    ord_rows = [dict(zip(ord_header, r)) for r in ords[1:]]

    labs = list(wb["检验"].iter_rows(values_only=True))
    lab_header = [str(h) for h in labs[0]]
    lab_rows = [dict(zip(lab_header, r)) for r in labs[1:]]

    vit = list(wb["体征"].iter_rows(values_only=True))
    vit_header = [str(h) for h in vit[0]]
    vit_rows = [dict(zip(vit_header, r)) for r in vit[1:]]
    wb.close()

    r = {"n_patients": len(adm_rows), "per_patient": {}, "frequency_codebook": {},
         "drug_orders_by_name": {}, "lab_analysis": {}, "vitals_analysis": {}}

    # 每患者概览
    for a in adm_rows:
        pid = str(a["PATIENT_ID"]).strip()
        r["per_patient"][pid] = {
            "age": a["AGE"], "sex": a["SEX"], "in": a["IN_DATE"], "out": a["DIS_DATE"],
            "days": a["DAY_TOTAL"], "main_dx": a["MAIN_DIAGNOSIS_NAME"],
            "other_dx": a["OTHER_DIAGNOSIS_NAME"], "fee": a["FEE_TOTAL"],
        }

    # 频次编码表
    freq_counter = Counter(str(o["FREQUENCY"]) for o in ord_rows if o.get("DRUG_FLAG") == 1)
    r["frequency_codebook"]["drug_orders"] = dict(freq_counter)

    # 药品医嘱按药名
    drug_orders = defaultdict(list)
    for o in ord_rows:
        if o.get("DRUG_FLAG") == 1:
            drug_orders[str(o["DRUG_NAME"])].append(o)
    r["drug_orders_by_name"] = {k: len(v) for k, v in sorted(drug_orders.items(), key=lambda x: -len(x[1]))}

    # 重点: TDM 目标药 (阿立哌唑/奥氮平/氯氮平/帕利哌酮/氯硝西泮) 医嘱详情
    target_orders = []
    for name, lst in drug_orders.items():
        if any(t in str(name) for t in ["阿立哌唑", "奥氮平", "氯氮平", "帕利哌酮", "氯硝西泮"]):
            for o in lst:
                target_orders.append({
                    "name": o["DRUG_NAME"], "dosage": o["DOSAGE"], "unit": o["DOSAGE_UNIT"],
                    "freq": o["FREQUENCY"], "way": o["MEDICATION_WAY"], "spec": o["DRUG_SPEC"],
                    "start": o["START_DATETIME"], "end": o["END_DATETIME"],
                    "state": o["ORDER_STATE"], "amount": o["AMOUNT"],
                })
    r["target_drug_orders"] = target_orders

    # 检验逐条: 时间(仅日期!)、目的、项目、结果、范围、患者
    lab_detail = []
    for l in lab_rows:
        lab_detail.append({
            "pid": str(l["PATIENT_ID"]).strip(), "purpose": l["TEST_PURPOSE"],
            "apply": l["APPLY_DATE"], "test": l["TEST_DATE"], "sample_no": l["SAMPLENO"],
            "project": l["PROJECT_NAME"], "result": l["TEST_RESULT"], "ref": l["REFFR_SCOPE"],
        })
    r["lab_analysis"]["detail"] = lab_detail
    r["lab_analysis"]["n_lab_rows"] = len(lab_detail)
    r["lab_analysis"]["has_sampling_time"] = any(
        re.search(r"\d{1,2}:\d{2}", str(l["TEST_DATE"])) for l in lab_rows)
    r["lab_analysis"]["has_sampling_time_note"] = "TEST_DATE 仅为日期,无采血时刻"

    # 派生量: 每患者每药物浓度
    by_patient_drug = defaultdict(list)
    for l in lab_detail:
        by_patient_drug[(l["pid"], l["project"])].append(l["result"])
    r["lab_analysis"]["concentrations_by_patient_project"] = {
        f"{k[0]}|{k[1]}": v for k, v in sorted(by_patient_drug.items())}

    # 达标率(用检验科 REFFR_SCOPE 判定)
    r["lab_analysis"]["range_achievement_by_reffr"] = {}
    for l in lab_detail:
        ref = l["ref"]
        if ref:
            m = re.match(r"([\d.]+)-([\d.]+)", str(ref))
            if m:
                lo, hi = float(m.group(1)), float(m.group(2))
                val = float(l["result"])
                key = (l["pid"], l["project"])
                r["lab_analysis"]["range_achievement_by_reffr"][
                    f"{l['pid']}|{l['project']}|{l['sample_no']}"] = {
                    "value": val, "ref": ref, "in_range": lo <= val <= hi,
                    "below": val < lo, "above": val > hi}

    # 活性部分: 阿立哌唑+脱氢阿立哌唑 (同一采样批次内)
    r["lab_analysis"]["active_moiety"] = {}
    by_sample = defaultdict(dict)
    for l in lab_detail:
        by_sample[l["sample_no"]][l["project"]] = float(l["result"])
    for sn, d in by_sample.items():
        if "阿立哌唑" in d and "脱氢阿立哌唑" in d:
            r["lab_analysis"]["active_moiety"][sn] = {
                "ari": d["阿立哌唑"], "dehydro": d["脱氢阿立哌唑"],
                "active_total": d["阿立哌唑"] + d["脱氢阿立哌唑"]}

    # 每日剂量可计算性: 目标药医嘱的频次是否可映射
    daily_dose_ok = 0
    daily_dose_map_fail = []
    for o in target_orders:
        f = o["freq"]
        if f in FREQ_DAILY_LO and FREQ_DAILY_LO[f] is not None:
            daily_dose_ok += 1
        else:
            daily_dose_map_fail.append({"name": o["name"], "dosage": o["dosage"], "freq": f})
    r["target_drug_orders_daily_dose"] = {
        "mappable": daily_dose_ok, "unmappable": daily_dose_map_fail}

    # 体征: 类型、血压解析
    vit_types = Counter(str(v["SIGN_TYPE"]) for v in vit_rows)
    r["vitals_analysis"]["types"] = dict(vit_types)
    bp_rows = [v for v in vit_rows if str(v["SIGN_TYPE"]) == "血压"]
    parsed_bp = []
    for v in bp_rows:
        m = re.match(r"(\d+)/(\d+)", str(v["RECORD_CONTENT"]))
        if m:
            parsed_bp.append({"pid": str(v["PATIENT_ID"]).strip(), "date": v["RECORD_DATE"],
                              "sys": int(m.group(1)), "dia": int(m.group(2))})
    r["vitals_analysis"]["n_bp"] = len(bp_rows)
    r["vitals_analysis"]["n_bp_parsed"] = len(parsed_bp)
    r["vitals_analysis"]["bp_sample"] = parsed_bp[:10]

    # 体重/身高可得性检查 (TDM 剂量标准化需要 mg/kg)
    r["vitals_analysis"]["has_weight_or_height"] = any(
        "体重" in str(v["SIGN_TYPE"]) or "身高" in str(v["SIGN_TYPE"]) for v in vit_rows)

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(r, f, ensure_ascii=False, indent=1, default=str)
    print(json.dumps({k: v for k, v in r.items() if k != "target_drug_orders"},
                     ensure_ascii=False, indent=1, default=str)[:6000])
    print("\n--- target orders (first 30) ---")
    for o in target_orders[:30]:
        print(o)


if __name__ == "__main__":
    main()

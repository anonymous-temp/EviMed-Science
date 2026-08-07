#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tdm_profile.py — 住院 TDM 数据集逐列剖析脚本
数据源: /workspace/20260803TDM.xlsx (5 张表: 病案首页/医嘱记录/检验/诊断记录/体征)
保留脚本以便复现;只读,不改写源数据。
"""
import json
import sys
from collections import Counter, OrderedDict
from datetime import datetime

import openpyxl

SRC = "/workspace/20260803TDM.xlsx"
OUT = "/workspace/tdm_profile_result.json"


def sheet_to_rows(ws, max_rows=None):
    """把 worksheet 转成 dict 列表(按首行表头)。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for r in rows[1:]:
        d = OrderedDict()
        for i, v in enumerate(r):
            if i < len(header):
                d[header[i]] = v
        data.append(d)
        if max_rows and len(data) >= max_rows:
            break
    return header, data


def analyze_column(name, values, sample_vals=None):
    """单列剖析: 类型、缺失率、唯一值、极值、示例。"""
    n = len(values)
    non_null = [v for v in values if v is not None and v != ""]
    n_missing = n - len(non_null)
    missing_rate = round(n_missing / n, 4) if n else None
    types = Counter(type(v).__name__ for v in non_null)
    uniq = len(set(str(v) for v in non_null)) if non_null else 0
    numeric = []
    for v in non_null:
        try:
            numeric.append(float(v))
        except (TypeError, ValueError):
            pass
    num_stat = None
    if numeric:
        num_stat = {
            "n_numeric": len(numeric),
            "min": round(min(numeric), 6),
            "max": round(max(numeric), 6),
            "mean": round(sum(numeric) / len(numeric), 6),
        }
    top = Counter(str(v) for v in non_null).most_common(8)
    return {
        "column": name,
        "n": n,
        "n_non_null": len(non_null),
        "missing_rate": missing_rate,
        "types": dict(types),
        "n_unique": uniq,
        "numeric_stats": num_stat,
        "top_values": top,
        "sample": sample_vals or non_null[:5],
    }


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    result = {"source": SRC, "sheets": {}}
    print(f"sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        header, rows = sheet_to_rows(ws)
        n_rows = ws.max_row - 1 if ws.max_row else 0
        print(f"\n=== Sheet: {sn} | header={len(header)} | rows(actual max_row-1)={n_rows} | parsed={len(rows)} ===")
        print("columns:", header)
        col_profiles = []
        if rows:
            # 转置: 每列取值
            by_col = {h: [] for h in header}
            for r in rows:
                for h in header:
                    by_col[h].append(r.get(h))
            for h in header:
                p = analyze_column(h, by_col[h])
                col_profiles.append(p)
                print(f"  - {h}: miss={p['missing_rate']} uniq={p['n_unique']} types={p['types']} "
                      f"num={p['numeric_stats']} top={p['top_values'][:4]}")
        result["sheets"][sn] = {
            "header": header,
            "n_columns": len(header),
            "n_rows": n_rows,
            "n_parsed": len(rows),
            "columns": col_profiles,
        }
    wb.close()
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=1, default=str)
    print(f"\nwrote {OUT}")


if __name__ == "__main__":
    main()
